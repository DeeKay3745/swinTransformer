"""
Marathi Varnamala Auto-Segmenter v3
====================================
- Detects speech segments from noisy WAV recordings
- Uses Whisper ASR to RECOGNIZE what character was spoken
- Works even if characters are out of order, repeated, or skipped
- Exports labeled WAV files + Excel report
 
Requirements:
    pip install pydub openpyxl openai-whisper torch
 
    # If whisper is slow, use the smaller model:
    # --model tiny  (fastest, less accurate)
    # --model base  (default, good balance)
    # --model small (slower, more accurate)
 
Usage:
    python segment_varnamala.py -i patient.wav --preview
    python segment_varnamala.py -i patient.wav -o patient_segments/
    python segment_varnamala.py -i patient.wav -o patient_segments/ --model small
"""
 
import os
import re
import argparse
import tempfile
from collections import Counter
from pydub import AudioSegment
from pydub.silence import detect_nonsilent
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
 
# ─────────────────────────────────────────────
#  Marathi Varnamala - Known characters
# ─────────────────────────────────────────────
VOWELS = ["अ", "आ", "ॲ", "ऑ", "इ", "ई", "उ", "ऊ", "ऋ", "ऌ", "ए", "ऐ", "ओ", "औ", "अं", "अः"]
 
CONSONANTS = [
    "क्", "ख्", "ग्", "घ्", "ङ्",
    "च्", "छ्", "ज्", "झ्", "ञ्",
    "ट्", "ठ्", "ड्", "ढ्", "ण्",
    "त्", "थ्", "द्", "ध्", "न्",
    "प",  "फ्", "ब्", "भ्", "म्",
    "य्", "र्", "ल्", "व्", "श्", "ष्", "स्", "ह्", "ळ्",
]
 
ALL_CHARS = VOWELS + CONSONANTS
VOWEL_SET = set(VOWELS)
CONSONANT_SET = set(CONSONANTS)
 
# Extended matching: map common Whisper outputs to actual characters
# Whisper sometimes outputs the full consonant (e.g., "क" instead of "क्")
CHAR_ALIASES = {}
for c in CONSONANTS:
    base = c.rstrip("्")
    if base != c:
        CHAR_ALIASES[base] = c
# Also add direct matches
for c in ALL_CHARS:
    CHAR_ALIASES[c] = c
 
 
def ms_to_timestamp(ms):
    secs = ms / 1000
    mins = int(secs // 60)
    secs = secs % 60
    return f"{mins:02d}:{secs:06.3f}"
 
 
def ms_to_seconds(ms):
    return round(ms / 1000, 3)
 
 
def match_to_known_char(whisper_text):
    """
    Try to match Whisper's recognized text to a known Marathi character.
    Returns (matched_char, confidence) or (None, 'unmatched').
    """
    text = whisper_text.strip()
 
    # Direct match
    if text in CHAR_ALIASES:
        return CHAR_ALIASES[text], "exact"
 
    # Check if any known character is contained in the text
    # (Whisper sometimes adds extra words around the character)
    for char in ALL_CHARS:
        if char in text:
            return char, "partial"
 
    # Check aliases in text
    for alias, char in CHAR_ALIASES.items():
        if alias in text:
            return char, "partial"
 
    # No match found
    return None, "unmatched"
 
 
def sanitize_filename(text):
    """Make text safe for filenames."""
    # Remove characters that are problematic in filenames
    safe = re.sub(r'[\\/:*?"<>|]', '', text)
    safe = safe.strip()
    return safe if safe else "unknown"
 
 
def load_whisper_model(model_name="base"):
    """Load Whisper model with Marathi language."""
    print(f"\n  Loading Whisper model '{model_name}'...")
    print(f"  (First run downloads the model, may take a minute)\n")
    import whisper
    model = whisper.load_model(model_name)
    return model
 
 
def recognize_segment(model, audio_segment, sample_rate=16000):
    """
    Run Whisper on a pydub AudioSegment, return recognized text.
    """
    import whisper
    import numpy as np
 
    # Convert to mono 16kHz for Whisper
    audio_16k = audio_segment.set_channels(1).set_frame_rate(sample_rate)
    samples = np.array(audio_16k.get_array_of_samples(), dtype=np.float32) / 32768.0
 
    # Pad if too short (Whisper needs at least ~0.1s)
    if len(samples) < sample_rate * 0.1:
        samples = np.pad(samples, (0, int(sample_rate * 0.1) - len(samples)))
 
    result = model.transcribe(
        samples,
        language="mr",  # Marathi
        task="transcribe",
        fp16=False,
        temperature=0.0,
        no_speech_threshold=0.3,
    )
 
    return result["text"].strip()
 
 
def create_excel_report(segments_info, output_dir, input_filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Segmentation Report"
 
    # Styles
    title_font = Font(name="Arial", size=14, bold=True, color="1F4E79")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    vowel_fill = PatternFill("solid", fgColor="E2EFDA")
    consonant_fill = PatternFill("solid", fgColor="D6E4F0")
    unmatched_fill = PatternFill("solid", fgColor="FCE4EC")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center")
 
    # Title
    ws.merge_cells("A1:I1")
    ws["A1"] = "Marathi Varnamala Segmentation Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")
 
    ws.merge_cells("A2:I2")
    ws["A2"] = f"Source: {input_filename}"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")
 
    # Headers
    headers = [
        "#", "Detected Character", "Whisper Raw Output", "Match Type",
        "Type", "Start Time", "End Time", "Duration (sec)", "Saved As"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
 
    # Data rows
    for i, info in enumerate(segments_info):
        row = i + 5
 
        matched = info["matched_char"]
        if matched and matched in VOWEL_SET:
            char_type = "Vowel (स्वर)"
            row_fill = vowel_fill
        elif matched and (matched in CONSONANT_SET or matched in CHAR_ALIASES.values()):
            char_type = "Consonant (व्यंजन)"
            row_fill = consonant_fill
        else:
            char_type = "Unmatched"
            row_fill = unmatched_fill
 
        values = [
            i + 1,
            info["matched_char"] or "???",
            info["whisper_text"],
            info["match_type"],
            char_type,
            info["start_time"],
            info["end_time"],
            info["duration_sec"],
            info["filename"],
        ]
 
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = center
 
        # Larger font for detected character
        ws.cell(row=row, column=2).font = Font(name="Arial", size=12, bold=True)
 
    # Summary
    summary_row = len(segments_info) + 6
    ws.merge_cells(f"A{summary_row}:C{summary_row}")
    ws[f"A{summary_row}"] = "Summary"
    ws[f"A{summary_row}"].font = Font(name="Arial", size=12, bold=True, color="1F4E79")
 
    matched_chars = [s["matched_char"] for s in segments_info if s["matched_char"]]
    vowel_count = sum(1 for c in matched_chars if c in VOWEL_SET)
    consonant_count = sum(1 for c in matched_chars if c in CONSONANT_SET or c in CHAR_ALIASES.values())
    unmatched_count = sum(1 for s in segments_info if not s["matched_char"])
    duplicates = sum(v - 1 for v in Counter(matched_chars).values() if v > 1)
 
    # Characters NOT found
    found_set = set(matched_chars)
    missing = [c for c in ALL_CHARS if c not in found_set]
 
    summary_data = [
        ("Total Segments Detected", len(segments_info)),
        ("Matched to Known Characters", len(matched_chars)),
        ("Vowels Found", vowel_count),
        ("Consonants Found", consonant_count),
        ("Unmatched Segments", unmatched_count),
        ("Duplicate Detections", duplicates),
        ("Missing Characters", len(missing)),
    ]
 
    for j, (label, val) in enumerate(summary_data):
        r = summary_row + 1 + j
        ws.cell(row=r, column=1, value=label).font = Font(name="Arial", size=10)
        ws.cell(row=r, column=2, value=val).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=r, column=1).border = border
        ws.cell(row=r, column=2).border = border
 
    # Missing characters list
    if missing:
        missing_row = summary_row + len(summary_data) + 2
        ws.merge_cells(f"A{missing_row}:I{missing_row}")
        ws[f"A{missing_row}"] = f"Missing characters: {' '.join(missing)}"
        ws[f"A{missing_row}"].font = Font(name="Arial", size=11, color="FF0000")
 
    # Column widths
    for i, w in enumerate([6, 18, 25, 14, 20, 14, 14, 14, 35], 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
 
    excel_path = os.path.join(output_dir, "segmentation_report.xlsx")
    wb.save(excel_path)
    return excel_path
 
 
def segment_audio(input_path, output_dir, silence_thresh=-35, min_silence_len=300,
                  min_duration=150, padding=50, preview=False, model_name="base"):
 
    print(f"\n{'='*60}")
    print(f"  Marathi Varnamala Auto-Segmenter v3")
    print(f"  (with Whisper ASR — order-independent)")
    print(f"{'='*60}")
    print(f"  Input:           {input_path}")
    print(f"  Whisper model:   {model_name}")
    print(f"  Silence thresh:  {silence_thresh} dBFS")
    print(f"  Min silence gap: {min_silence_len} ms")
    print(f"  Min duration:    {min_duration} ms")
    print(f"  Padding:         {padding} ms")
    print(f"  Known chars:     {len(ALL_CHARS)}")
    print(f"{'='*60}\n")
 
    # Load audio
    print("Loading audio...")
    audio = AudioSegment.from_wav(input_path)
    print(f"  Duration: {ms_to_timestamp(len(audio))} | "
          f"Sample rate: {audio.frame_rate} Hz | "
          f"Channels: {audio.channels}")
 
    audio_mono = audio.set_channels(1) if audio.channels > 1 else audio
 
    # Detect speech
    print("\nDetecting speech segments...")
    nonsilent_ranges = detect_nonsilent(
        audio_mono,
        min_silence_len=min_silence_len,
        silence_thresh=silence_thresh,
        seek_step=10
    )
 
    filtered = [(s, e) for s, e in nonsilent_ranges if (e - s) >= min_duration]
 
    print(f"  Raw segments:    {len(nonsilent_ranges)}")
    print(f"  After filtering: {len(filtered)}")
 
    # Load Whisper
    model = load_whisper_model(model_name)
 
    # Process each segment
    print(f"  Recognizing each segment with Whisper (language: Marathi)...\n")
    print(f"{'─'*85}")
    print(f"  {'#':<5} {'Detected':<10} {'Whisper Output':<25} {'Match':<10} {'Start':<12} {'End':<12} {'Dur'}")
    print(f"{'─'*85}")
 
    segments_info = []
    char_count = Counter()  # Track duplicates for unique filenames
 
    for i, (start, end) in enumerate(filtered):
        # Extract segment with padding
        seg_start = max(0, start - padding)
        seg_end = min(len(audio), end + padding)
        segment = audio[seg_start:seg_end]
 
        # Recognize with Whisper
        whisper_text = recognize_segment(model, segment)
        matched_char, match_type = match_to_known_char(whisper_text)
 
        # Generate filename
        if matched_char:
            char_count[matched_char] += 1
            count = char_count[matched_char]
            safe_char = sanitize_filename(matched_char)
 
            if matched_char in VOWEL_SET:
                folder = "vowels_swar"
            else:
                folder = "consonants_vyanjan"
 
            if count > 1:
                filename = f"{folder}/{safe_char}_repeat{count}.wav"
            else:
                filename = f"{folder}/{safe_char}.wav"
        else:
            safe_text = sanitize_filename(whisper_text[:20]) if whisper_text else f"segment_{i+1}"
            filename = f"unmatched/{i+1:03d}_{safe_text}.wav"
 
        duration = end - start
        info = {
            "label": matched_char or whisper_text or "???",
            "matched_char": matched_char,
            "whisper_text": whisper_text,
            "match_type": match_type,
            "start_ms": start,
            "end_ms": end,
            "start_time": ms_to_timestamp(start),
            "end_time": ms_to_timestamp(end),
            "duration_sec": ms_to_seconds(duration),
            "filename": filename,
            "segment_audio": segment,
        }
        segments_info.append(info)
 
        display_char = matched_char or "???"
        print(f"  {i+1:<5} {display_char:<10} {whisper_text:<25} {match_type:<10} "
              f"{ms_to_timestamp(start):<12} {ms_to_timestamp(end):<12} {duration}ms")
 
    # Stats
    matched = [s for s in segments_info if s["matched_char"]]
    unmatched = [s for s in segments_info if not s["matched_char"]]
    found_set = set(s["matched_char"] for s in matched)
    missing = [c for c in ALL_CHARS if c not in found_set]
 
    print(f"\n{'─'*85}")
    print(f"  ✅ Matched:    {len(matched)}/{len(segments_info)} segments")
    print(f"  ❓ Unmatched:  {len(unmatched)} segments")
    print(f"  📋 Found:      {len(found_set)}/{len(ALL_CHARS)} unique characters")
    if missing:
        print(f"  ⚠️  Missing:   {' '.join(missing)}")
 
    if preview:
        print(f"\n  🔍 Preview mode — no files exported.")
        print(f"  Remove --preview to export.\n")
        return
 
    # ── Export ──
    os.makedirs(output_dir, exist_ok=True)
    for folder in ["vowels_swar", "consonants_vyanjan", "unmatched"]:
        os.makedirs(os.path.join(output_dir, folder), exist_ok=True)
 
    print(f"\nExporting audio segments...")
    for info in segments_info:
        filepath = os.path.join(output_dir, info["filename"])
        info["segment_audio"].export(filepath, format="wav")
 
    # Remove audio objects before Excel (not serializable)
    for info in segments_info:
        del info["segment_audio"]
 
    print(f"  ✅ Exported {len(segments_info)} audio files!")
 
    # Excel report
    print(f"Generating Excel report...")
    excel_path = create_excel_report(segments_info, output_dir, os.path.basename(input_path))
    print(f"  ✅ Report saved: {excel_path}")
 
    # Final summary
    print(f"\n{'='*60}")
    print(f"  📁 Output: {output_dir}/")
    vowel_count = sum(1 for s in segments_info if s["matched_char"] and s["matched_char"] in VOWEL_SET)
    cons_count = sum(1 for s in segments_info if s["matched_char"] and s["matched_char"] not in VOWEL_SET and s["matched_char"])
    unmatched_count = len(unmatched)
    print(f"     ├── vowels_swar/           ({vowel_count} files)")
    print(f"     ├── consonants_vyanjan/    ({cons_count} files)")
    if unmatched_count:
        print(f"     ├── unmatched/             ({unmatched_count} files)")
    print(f"     └── segmentation_report.xlsx")
    if missing:
        print(f"\n  ⚠️  Characters NOT found in audio:")
        print(f"     {' '.join(missing)}")
    print(f"{'='*60}\n")
 
 
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Auto-segment Marathi Varnamala with Whisper recognition",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
HOW IT WORKS:
─────────────
  1. Splits audio by detecting silence gaps
  2. Runs Whisper ASR on each segment (Marathi language)
  3. Matches recognized text to known varnamala characters
  4. Names files by the DETECTED character (not position)
  5. Handles repeats, skips, and out-of-order speech
 
WHISPER MODELS (speed vs accuracy):
  --model tiny    → Fastest, least accurate
  --model base    → Default, good balance
  --model small   → Slower, more accurate
  --model medium  → Much slower, best for difficult audio
 
TUNING SILENCE DETECTION:
  Too many segments?  → --silence_thresh -30 or --min_duration 250
  Too few segments?   → --silence_thresh -45 or --min_silence 200
  Noisy audio?        → --silence_thresh -40 to -45
  Slow speaker?       → --min_silence 500
 
EXAMPLES:
  python segment_varnamala.py -i patient.wav --preview
  python segment_varnamala.py -i patient.wav -o patient_01/ --model small
  python segment_varnamala.py -i patient.wav -o patient_01/ -t -42 -s 400
        """,
    )
 
    parser.add_argument("--input", "-i", required=True, help="Input WAV file")
    parser.add_argument("--output_dir", "-o", default="output_segments", help="Output directory")
    parser.add_argument("--model", "-m", default="base", choices=["tiny", "base", "small", "medium"],
                        help="Whisper model size (default: base)")
    parser.add_argument("--silence_thresh", "-t", type=int, default=-35, help="Silence threshold dBFS")
    parser.add_argument("--min_silence", "-s", type=int, default=300, help="Min silence gap ms")
    parser.add_argument("--min_duration", "-d", type=int, default=150, help="Min segment duration ms")
    parser.add_argument("--padding", "-p", type=int, default=50, help="Padding ms")
    parser.add_argument("--preview", action="store_true", help="Preview only")
 
    args = parser.parse_args()
    segment_audio(
        input_path=args.input,
        output_dir=args.output_dir,
        silence_thresh=args.silence_thresh,
        min_silence_len=args.min_silence,
        min_duration=args.min_duration,
        padding=args.padding,
        preview=args.preview,
        model_name=args.model,
    )