"""
Dysarthric Speech Audio Trimmer  v5
=====================================
- Characters & Words  : best-match anywhere in audio (working fine)
- Sentences & Paragraph: anchor on DISTINCTIVE words (long, rare, unique to that sentence)
                         then take the time window around those anchors.
                         Common stop-words ("the","a","is") are IGNORED for anchoring.

INSTALL:
    pip install openai-whisper stable-ts pydub openpyxl

USAGE:
    python audio_trimmer.py --audio file.wav --method stable --model medium
    python audio_trimmer.py --audio file.wav --save_json   # debug: see what Whisper heard
    python audio_trimmer.py --audio file.wav --no_trim     # Excel only
"""

import argparse, os, sys, json, re, warnings
warnings.filterwarnings("ignore")


# ═══════════════════════════════════════════════════════════════════════════
#  STOP WORDS  (ignored when anchoring sentences – too common to be useful)
# ═══════════════════════════════════════════════════════════════════════════
STOP = {
    "the","a","an","is","it","in","of","to","and","or","be","are","was",
    "were","has","have","had","that","this","as","at","by","for","on",
    "its","with","from","but","not","no","so","if","he","she","they",
    "we","you","i","me","my","his","her","our","their","been","can",
    "will","would","could","should","may","might","do","did","does","up",
    "out","then","than","when","which","who","what","how","there","here",
    "all","one","two","into","upon","upon","upon","some","more","such",
    "said","say","also","each","both","per","via","yet","now","over",
    "end","say","made","make","give","take","find","found"
}


# ═══════════════════════════════════════════════════════════════════════════
#  EXPECTED CONTENT
# ═══════════════════════════════════════════════════════════════════════════
ALPHABETS = list("ABCDEFGHIJKLMNOPQRSTUVWXYZ")

WORDS = [
    "One","three","four","five","seven","twelve","fifteen","twenty-nine",
    "Their","If","Alpha","Beta","Delta","Could","Adapt","Circular",
    "Composure","Footwork","Journalism","Python","Advice","Choice",
    "Employment","Immovable","Massage","Moisten","Tree","Knife",
    "Spoon","Banana","Monkey",
]

SENTENCES = [
    ("S1","Each untimely income loss coincided with the breakdown of a heating system part."),
    ("S2","Alice's ability to work without supervision is noteworthy."),
    ("S3","Special task forces rescue hostages from kidnappers."),
    ("S4","Laugh, dance, and sing if fortune smiles upon you."),
    ("S5","The same shelter could be built into an embankment or below ground level."),
]

# Paragraph split into individual sentences for better localisation
PARA_SENTENCES = [
    "When the sunlight strikes raindrops in the air they act as a prism and form a rainbow",
    "The rainbow is a division of white light into many beautiful colors",
    "These take the shape of a long round arch with its path high above and its two ends apparently beyond the horizon",
    "There is according to legend a boiling pot of gold at one end",
    "People look but no one ever finds it",
    "When a man looks for something beyond his reach his friends say he is looking for the pot of gold at the end of the rainbow",
    "Throughout the centuries people have explained the rainbow in various ways",
    "Some have accepted it as a miracle without physical explanation",
    "To the Hebrews it was a token that there would be no more universal floods",
    "The Greeks used to imagine that it was a sign from the gods to foretell war or heavy rain",
    "The Norsemen considered the rainbow as a bridge over which the gods passed from earth to their home in the sky",
    "Others have tried to explain the phenomenon physically",
    "Aristotle thought that the rainbow was caused by reflection of the suns rays by the rain",
    "Since then physicists have found that it is not reflection but refraction by the raindrops which causes the rainbows",
    "Many complicated ideas about the rainbow have been formed",
    "The difference in the rainbow depends considerably upon the size of the drops and the width of the colored band increases as the size of the drops increases",
    "The actual primary rainbow observed is said to be the effect of super imposition of a number of bows",
    "If the red of the second bow falls upon the green of the first the result is to give a bow with an abnormally wide yellow band since red and green light when mixed form yellow",
    "This is a very common type of bow one showing mainly red and yellow with little or no green or blue",
]

PARAGRAPH_FULL = " ".join(PARA_SENTENCES)


# ═══════════════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════════════
def norm(s):
    return re.sub(r"[^a-z]", "", s.lower())

def norm_words(text):
    return [norm(w) for w in text.split() if norm(w)]

def distinctive_words(text, min_len=5):
    """Return normalised words that are long enough and not stop-words."""
    return [w for w in norm_words(text) if len(w) >= min_len and w not in STOP]

def fmt_time(seconds):
    if seconds is None:
        return "—"
    h  = int(seconds // 3600)
    m  = int((seconds % 3600) // 60)
    s  = seconds % 60
    return f"{h:02d}:{m:02d}:{s:06.3f}"

def fuzzy_match(expected, detected):
    """Score 0..1 for how well two normalised words match."""
    if expected == detected:                           return 1.0
    ml = min(len(expected), len(detected))
    if ml >= 4 and expected[:4] == detected[:4]:       return 0.8
    if len(expected) >= 4 and (expected in detected
                                or detected in expected): return 0.6
    if ml >= 3 and expected[:3] == detected[:3]:       return 0.5
    return 0.0


# ═══════════════════════════════════════════════════════════════════════════
#  STEP 1 – TRANSCRIBE
# ═══════════════════════════════════════════════════════════════════════════
def transcribe_stable(audio_path, model_size):
    import stable_whisper
    print(f"\n[1/4] Loading stable-ts '{model_size}' model ...")
    model = stable_whisper.load_model(model_size)
    print(f"[2/4] Transcribing ...")
    result = model.transcribe(
        audio_path, language="en",
        regroup=True, suppress_silence=True,
        word_timestamps=True, vad=True,
    )
    words = []
    for seg in result.segments:
        for w in seg.words:
            t = norm(w.word)
            if t:
                words.append({"word": t,
                               "start": round(w.start, 3),
                               "end":   round(w.end,   3)})
    print(f"       -> {len(words)} words detected")
    return words


def transcribe_whisper(audio_path, model_size):
    import whisper
    print(f"\n[1/4] Loading Whisper '{model_size}' model ...")
    model = whisper.load_model(model_size)
    print(f"[2/4] Transcribing ...")
    result = model.transcribe(
        audio_path, language="en", word_timestamps=True,
        condition_on_previous_text=False, temperature=0.0,
        no_speech_threshold=0.35,
    )
    words = []
    for seg in result.get("segments", []):
        for w in seg.get("words", []):
            t = norm(w["word"])
            if t:
                words.append({"word": t,
                               "start": round(w["start"], 3),
                               "end":   round(w["end"],   3)})
    print(f"       -> {len(words)} words detected")
    return words


# ═══════════════════════════════════════════════════════════════════════════
#  STEP 2A – SINGLE WORD MATCH  (for alphabets & single-token words)
# ═══════════════════════════════════════════════════════════════════════════
def match_single(expected_norm, detected_words, used):
    best_idx, best_score = None, 0.0
    for i, dw in enumerate(detected_words):
        if i in used:
            continue
        score = fuzzy_match(expected_norm, dw["word"])
        if score > best_score:
            best_score, best_idx = score, i
    if best_idx is not None and best_score >= 0.5:
        return best_idx, detected_words[best_idx]["start"], detected_words[best_idx]["end"]
    return None


# ═══════════════════════════════════════════════════════════════════════════
#  STEP 2B – SENTENCE / PARAGRAPH MATCH
#
#  Algorithm:
#  1. Find ALL detected-word positions that match any DISTINCTIVE word of
#     the sentence (long, non-stop-word).
#  2. For each candidate anchor position, score a ±expansion window:
#     count how many TOTAL sentence words (including common ones) appear
#     within that window.
#  3. Pick the window with the highest total coverage.
#  4. Return (start, end) of the winning window.
# ═══════════════════════════════════════════════════════════════════════════
def match_sentence(sentence_text, detected_words, used,
                   max_window_s=45.0, min_match_ratio=0.30):
    """
    Returns (start_s, end_s, matched_count, total_count, ratio_pct)
    or None if no good match found.
    """
    all_exp   = norm_words(sentence_text)          # every word
    key_exp   = distinctive_words(sentence_text)   # long/distinctive only

    if not all_exp:
        return None

    # ── 1. Collect candidate anchor positions from DISTINCTIVE words only ──
    anchors = []   # (time_start, det_index)
    for exp_w in key_exp:
        for i, dw in enumerate(detected_words):
            if i in used:
                continue
            if fuzzy_match(exp_w, dw["word"]) >= 0.5:
                anchors.append((dw["start"], i))

    # Fall back to ALL words if no distinctive matches
    if not anchors:
        for exp_w in all_exp:
            if exp_w in STOP:
                continue
            for i, dw in enumerate(detected_words):
                if i in used:
                    continue
                if fuzzy_match(exp_w, dw["word"]) >= 0.8:   # stricter for common words
                    anchors.append((dw["start"], i))

    if not anchors:
        return None

    # ── 2. For each anchor, build a window and count ALL word coverage ──
    best_score   = 0
    best_result  = None

    seen_anchors = set()
    for anchor_t, anchor_i in anchors:
        # deduplicate nearby anchors (within 1s)
        bucket = int(anchor_t)
        if bucket in seen_anchors:
            continue
        seen_anchors.add(bucket)

        # Collect every detected word (not used) within the window
        window_words = []
        for i, dw in enumerate(detected_words):
            if i in used:
                continue
            if anchor_t - 2.0 <= dw["start"] <= anchor_t + max_window_s:
                window_words.append((i, dw["word"], dw["start"], dw["end"]))

        if not window_words:
            continue

        # Score: how many expected words are covered by this window?
        covered = set()
        used_in_window = []
        for exp_w in all_exp:
            for (wi, wword, wstart, wend) in window_words:
                if wi in used_in_window:
                    continue
                if fuzzy_match(exp_w, wword) >= 0.5:
                    covered.add(exp_w)
                    used_in_window.append(wi)
                    break

        score = len(covered) / len(all_exp)
        if score > best_score:
            best_score  = score
            # window boundaries = first..last matched word in this window
            matched_window = [w for w in window_words if w[0] in used_in_window]
            if matched_window:
                start_t = min(w[2] for w in matched_window)
                end_t   = max(w[3] for w in matched_window)
                best_result = (start_t, end_t,
                               [w[0] for w in matched_window],
                               len(covered), len(all_exp),
                               round(score * 100))

    if best_result is None:
        return None

    start_t, end_t, indices, matched, total, ratio = best_result
    if ratio < min_match_ratio * 100:
        return None

    return start_t, end_t, indices, matched, total, ratio


def match_paragraph(detected_words, used):
    """
    Match the full paragraph by finding each paragraph sentence individually,
    then span from the earliest to the latest matched sentence.
    """
    all_starts, all_ends, all_indices = [], [], []
    total_matched = 0
    total_words   = len(norm_words(PARAGRAPH_FULL))
    matched_sents = 0

    for sent in PARA_SENTENCES:
        result = match_sentence(sent, detected_words, used,
                                max_window_s=60.0, min_match_ratio=0.20)
        if result:
            s, e, idxs, matched, total, ratio = result
            all_starts.append(s)
            all_ends.append(e)
            all_indices.extend(idxs)
            total_matched += matched
            matched_sents += 1
            for i in idxs:
                used.add(i)

    if not all_starts:
        return None

    start_t = min(all_starts)
    end_t   = max(all_ends)
    ratio   = round(total_matched / total_words * 100)
    note    = f"{matched_sents}/{len(PARA_SENTENCES)} sentences, {total_matched}/{total_words} words ({ratio}%)"
    return start_t, end_t, list(set(all_indices)), note


# ═══════════════════════════════════════════════════════════════════════════
#  STEP 2 – BUILD ALL SEGMENTS
# ═══════════════════════════════════════════════════════════════════════════
def build_segments(detected_words):
    used = set()
    segments = []
    sid = 1

    print(f"\n[3/4] Matching segments ...")

    # ── ALPHABETS ─────────────────────────────────────────────────────────
    for letter in ALPHABETS:
        r = match_single(norm(letter), detected_words, used)
        if r:
            idx, s, e = r
            used.add(idx)
            segments.append(_seg(sid, "Alphabet", letter, s, e, ""))
        else:
            segments.append(_empty(sid, "Alphabet", letter, "Not detected"))
        sid += 1

    # ── WORDS ─────────────────────────────────────────────────────────────
    for word in WORDS:
        nw = norm_words(word)
        if len(nw) == 1:
            r = match_single(nw[0], detected_words, used)
            if r:
                idx, s, e = r
                used.add(idx)
                segments.append(_seg(sid, "Word", word, s, e, ""))
            else:
                segments.append(_empty(sid, "Word", word, "Not detected"))
        else:
            # multi-token word (e.g. "twenty-nine") — use sentence matcher
            r = match_sentence(word, detected_words, used,
                               max_window_s=5.0, min_match_ratio=0.5)
            if r:
                s, e, idxs, matched, total, ratio = r
                for i in idxs: used.add(i)
                segments.append(_seg(sid, "Word", word, s, e,
                                     f"{matched}/{total} parts ({ratio}%)"))
            else:
                segments.append(_empty(sid, "Word", word, "Not detected"))
        sid += 1

    # ── SENTENCES ─────────────────────────────────────────────────────────
    for label, sentence in SENTENCES:
        r = match_sentence(sentence, detected_words, used,
                           max_window_s=45.0, min_match_ratio=0.30)
        if r:
            s, e, idxs, matched, total, ratio = r
            for i in idxs: used.add(i)
            segments.append(_seg(sid, f"Sentence {label}", sentence, s, e,
                                 f"{matched}/{total} words ({ratio}%)"))
        else:
            segments.append(_empty(sid, f"Sentence {label}", sentence, "Not detected"))
        sid += 1

    # ── PARAGRAPH ─────────────────────────────────────────────────────────
    r = match_paragraph(detected_words, used)
    if r:
        s, e, idxs, note = r
        segments.append(_seg(sid, "Paragraph", PARAGRAPH_FULL[:120] + "...", s, e, note))
    else:
        segments.append(_empty(sid, "Paragraph",
                               PARAGRAPH_FULL[:120] + "...", "Not detected"))

    detected_count = sum(1 for sg in segments if sg["Start_s"] is not None)
    print(f"       -> {detected_count}/{len(segments)} segments matched")
    return segments


def _seg(sid, stype, content, s, e, note):
    return {"Segment_ID": sid, "Type": stype, "Content": content,
            "Start_s": s, "End_s": e, "Note": note, "File": "—"}

def _empty(sid, stype, content, note):
    return {"Segment_ID": sid, "Type": stype, "Content": content,
            "Start_s": None, "End_s": None, "Note": note, "File": "—"}


# ═══════════════════════════════════════════════════════════════════════════
#  STEP 3 – TRIM AUDIO
# ═══════════════════════════════════════════════════════════════════════════
def trim_audio(audio_path, segments, output_dir, padding_ms=200):
    from pydub import AudioSegment
    audio    = AudioSegment.from_wav(audio_path)
    total_ms = len(audio)
    os.makedirs(output_dir, exist_ok=True)
    saved = 0
    for seg in segments:
        if seg["Start_s"] is None:
            continue
        s_ms = max(0,        int(seg["Start_s"] * 1000) - padding_ms)
        e_ms = min(total_ms, int(seg["End_s"]   * 1000) + padding_ms)
        clip = audio[s_ms:e_ms]
        safe  = re.sub(r"[^\w\-]", "_", seg["Content"][:35]).strip("_")
        fname = (f"{seg['Segment_ID']:04d}_"
                 f"{seg['Type'].replace(' ','_')}_{safe}.wav")
        fpath = os.path.join(output_dir, fname)
        clip.export(fpath, format="wav")
        seg["File"] = fpath
        saved += 1
    print(f"       -> {saved} clips saved to '{output_dir}/'")


# ═══════════════════════════════════════════════════════════════════════════
#  STEP 4 – EXCEL REPORT
# ═══════════════════════════════════════════════════════════════════════════
def export_excel(segments, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Audio Segments"

    THIN   = Side(style="thin", color="BBBBBB")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    TYPE_COLOR = {
        "Alphabet":  "FFF2CC",
        "Word":      "E2EFDA",
        "Sentence":  "DDEBF7",
        "Paragraph": "FCE4D6",
    }

    headers    = ["ID", "Type", "Content / Text Spoken",
                  "From (HH:MM:SS.mmm)", "To (HH:MM:SS.mmm)",
                  "From (s)", "To (s)", "Duration (ms)",
                  "Trimmed .wav File", "Match Info"]
    col_widths = [5, 15, 62, 22, 22, 11, 11, 14, 55, 28]

    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", start_color="1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 32

    for r, seg in enumerate(segments, start=2):
        tk   = seg["Type"].split()[0]
        fill = PatternFill("solid", start_color=TYPE_COLOR.get(tk, "FFFFFF"))
        dur  = (round((seg["End_s"] - seg["Start_s"]) * 1000)
                if seg["Start_s"] is not None else None)
        vals = [
            seg["Segment_ID"], seg["Type"], seg["Content"],
            fmt_time(seg["Start_s"]), fmt_time(seg["End_s"]),
            seg["Start_s"], seg["End_s"], dur,
            seg.get("File", "—"), seg.get("Note", ""),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = fill
            cell.border    = BORDER
            cell.alignment = Alignment(vertical="center",
                                       wrap_text=(c in (3, 9)))
            if c in (6, 7) and isinstance(v, float):
                cell.number_format = "0.000"
            if c == 8 and isinstance(v, int):
                cell.number_format = "#,##0"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["B"].width = 34
    ws2.column_dimensions["C"].width = 18

    def s2(row, col, value, bold=False, bg=None, fmt=None):
        cell = ws2.cell(row=row, column=col, value=value)
        cell.font = Font(name="Arial", bold=bold, size=11)
        if bg:
            cell.fill = PatternFill("solid", start_color=bg)
        if fmt:
            cell.number_format = fmt
        return cell

    total    = len(segments)
    detected = sum(1 for s in segments if s["Start_s"] is not None)

    s2(1, 2, "Dysarthric Speech Segmentation Report", bold=True).font = Font(
        name="Arial", bold=True, size=14, color="1F3864")
    s2(3, 2, "Total segments expected")
    s2(3, 3, total)
    s2(4, 2, "Segments detected",  bg="E2EFDA")
    s2(4, 3, detected,             bg="E2EFDA")
    s2(5, 2, "Segments NOT found", bg="FCE4D6")
    s2(5, 3, total - detected,     bg="FCE4D6")
    s2(6, 2, "Detection rate")
    s2(6, 3, "=C5/C4",             fmt="0.0%")
    s2(8, 2, "BREAKDOWN BY TYPE",  bold=True, bg="1F3864").font = Font(
        name="Arial", bold=True, color="FFFFFF", size=11)
    s2(8, 3, "Detected / Total",   bold=True, bg="1F3864").font = Font(
        name="Arial", bold=True, color="FFFFFF", size=11)

    type_data = {}
    for seg in segments:
        tk = seg["Type"].split()[0]
        type_data.setdefault(tk, [0, 0])
        type_data[tk][1] += 1
        if seg["Start_s"] is not None:
            type_data[tk][0] += 1

    for i, (tk, (det, tot)) in enumerate(type_data.items(), start=9):
        color = TYPE_COLOR.get(tk, "FFFFFF")
        s2(i, 2, tk,               bg=color)
        s2(i, 3, f"{det} / {tot}", bg=color)

    wb.save(output_path)
    print(f"\n       Excel saved -> '{output_path}'")


# ═══════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--audio",      required=True)
    parser.add_argument("--model",      default="medium",
                        choices=["tiny","base","small","medium",
                                 "large","large-v2","large-v3"])
    parser.add_argument("--method",     default="stable",
                        choices=["stable","whisper"])
    parser.add_argument("--output_dir", default="trimmed_segments")
    parser.add_argument("--excel",      default="audio_timestamps.xlsx")
    parser.add_argument("--padding_ms", type=int, default=200)
    parser.add_argument("--no_trim",    action="store_true")
    parser.add_argument("--save_json",  action="store_true",
                        help="Save every detected word+timestamp to JSON for debugging")
    args = parser.parse_args()

    if not os.path.exists(args.audio):
        print(f"ERROR: File not found - '{args.audio}'")
        sys.exit(1)

    # 1. Transcribe
    if args.method == "stable":
        words = transcribe_stable(args.audio, args.model)
    else:
        words = transcribe_whisper(args.audio, args.model)

    if args.save_json:
        jpath = args.excel.replace(".xlsx", "_words.json")
        with open(jpath, "w") as f:
            json.dump(words, f, indent=2)
        print(f"       Raw word list -> '{jpath}'")

    # 2. Match
    segments = build_segments(words)

    # 3. Trim audio
    if not args.no_trim:
        trim_audio(args.audio, segments, args.output_dir, args.padding_ms)

    # 4. Excel
    export_excel(segments, args.excel)

    detected = sum(1 for s in segments if s["Start_s"] is not None)
    print("\n===============================================")
    print(f"  Segments : {detected}/{len(segments)} detected")
    if not args.no_trim:
        print(f"  Clips    : {args.output_dir}/")
    print(f"  Excel    : {args.excel}")
    print("===============================================\n")


if __name__ == "__main__":
    main()