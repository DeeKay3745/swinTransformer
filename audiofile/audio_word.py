"""
Word Detector – Dysarthric Speech
===================================
Detects ONLY the target words from a .wav file and exports
From/To timestamps + trimmed clips to Excel.

INSTALL:
    pip install openai-whisper stable-ts pydub openpyxl

USAGE:
    python word_detector.py --audio your_file.wav --method stable --model medium
    python word_detector.py --audio your_file.wav --save_json
    python word_detector.py --audio your_file.wav --no_trim
"""

import argparse, os, sys, json, re, warnings
warnings.filterwarnings("ignore")


# ═══════════════════════════════════════════════════════════════════════════
#  TARGET WORDS  (in document order)
# ═══════════════════════════════════════════════════════════════════════════
TARGET_WORDS = [
    "One", "three", "four", "five", "seven", "twelve", "fifteen", "twenty-nine",
    "Their", "If",
    "Alpha", "Beta", "Delta", "Could", "Adapt", "Circular",
    "Composure", "Footwork", "Journalism", "Python", "Advice", "Choice",
    "Employment", "Immovable", "Massage", "Moisten",
    "Tree", "Knife", "Spoon", "Banana", "Monkey",
]


# ═══════════════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════════════
def norm(s):
    return re.sub(r"[^a-z]", "", s.lower())

def fmt_time(seconds):
    if seconds is None:
        return "—"
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    s = seconds % 60
    return f"{h:02d}:{m:02d}:{s:06.3f}"

def fuzzy_score(expected, detected):
    """Return match score 0..1 between two normalised words."""
    if expected == detected:
        return 1.0
    ml = min(len(expected), len(detected))
    if ml >= 4 and expected[:4] == detected[:4]:
        return 0.8
    if len(expected) >= 4 and (expected in detected or detected in expected):
        return 0.6
    if ml >= 3 and expected[:3] == detected[:3]:
        return 0.5
    return 0.0


# ═══════════════════════════════════════════════════════════════════════════
#  STEP 1 – TRANSCRIBE
# ═══════════════════════════════════════════════════════════════════════════
def transcribe_stable(audio_path, model_size):
    import stable_whisper
    print(f"\n[1/4] Loading stable-ts '{model_size}' model ...")
    model = stable_whisper.load_model(model_size)
    print(f"[2/4] Transcribing with word-level alignment ...")
    result = model.transcribe(
        audio_path,
        language="en",
        regroup=True,
        suppress_silence=True,
        word_timestamps=True,
        vad=True,
    )
    words = []
    for seg in result.segments:
        for w in seg.words:
            t = norm(w.word)
            if t:
                words.append({
                    "word":  t,
                    "start": round(w.start, 3),
                    "end":   round(w.end,   3),
                })
    print(f"       -> {len(words)} words detected in audio")
    return words


def transcribe_whisper(audio_path, model_size):
    import whisper
    print(f"\n[1/4] Loading Whisper '{model_size}' model ...")
    model = whisper.load_model(model_size)
    print(f"[2/4] Transcribing ...")
    result = model.transcribe(
        audio_path,
        language="en",
        word_timestamps=True,
        condition_on_previous_text=False,
        temperature=0.0,
        no_speech_threshold=0.35,
    )
    words = []
    for seg in result.get("segments", []):
        for w in seg.get("words", []):
            t = norm(w["word"])
            if t:
                words.append({
                    "word":  t,
                    "start": round(w["start"], 3),
                    "end":   round(w["end"],   3),
                })
    print(f"       -> {len(words)} words detected in audio")
    return words


# ═══════════════════════════════════════════════════════════════════════════
#  STEP 2 – MATCH TARGET WORDS
#  For multi-token words like "twenty-nine", find a cluster of parts.
#  For single-token words, find best fuzzy match anywhere in audio.
#  Each detected word can only be used ONCE (used set).
# ═══════════════════════════════════════════════════════════════════════════
def match_single(exp_norm, detected, used):
    """Best unused detected word matching exp_norm."""
    best_idx, best_score = None, 0.0
    for i, dw in enumerate(detected):
        if i in used:
            continue
        score = fuzzy_score(exp_norm, dw["word"])
        if score > best_score:
            best_score, best_idx = score, i
    if best_idx is not None and best_score >= 0.5:
        return best_idx, detected[best_idx]["start"], detected[best_idx]["end"]
    return None


def match_multi(parts_norm, detected, used, max_gap_s=4.0):
    """
    Find a time-clustered group of detected words matching all parts
    of a compound word (e.g. 'twenty' + 'nine').
    """
    # Find candidate positions for each part
    candidates = []
    for part in parts_norm:
        for i, dw in enumerate(detected):
            if i in used:
                continue
            if fuzzy_score(part, dw["word"]) >= 0.5:
                candidates.append((dw["start"], dw["end"], i, part))

    if not candidates:
        return None

    candidates.sort(key=lambda x: x[0])

    best, best_score = None, 0
    for anchor in candidates:
        anchor_t = anchor[0]
        window   = [c for c in candidates
                    if anchor_t <= c[0] <= anchor_t + max_gap_s]
        covered  = set(c[3] for c in window)
        if len(covered) > best_score:
            best_score = len(covered)
            best       = window

    if not best or best_score < len(parts_norm) * 0.5:
        return None

    start_t  = min(c[0] for c in best)
    end_t    = max(c[1] for c in best)
    indices  = [c[2] for c in best]
    ratio    = round(best_score / len(parts_norm) * 100)
    return start_t, end_t, indices, best_score, ratio


def build_segments(detected):
    """Match every target word and return segment list."""
    used = set()
    segments = []
    print(f"\n[3/4] Matching {len(TARGET_WORDS)} target words ...")

    for idx, word in enumerate(TARGET_WORDS, start=1):
        parts = [norm(p) for p in re.split(r"[-\s]", word) if norm(p)]

        if len(parts) == 1:
            r = match_single(parts[0], detected, used)
            if r:
                det_idx, start, end = r
                used.add(det_idx)
                segments.append({
                    "ID":      idx,
                    "Word":    word,
                    "Start_s": start,
                    "End_s":   end,
                    "Note":    "matched",
                    "File":    "—",
                })
            else:
                segments.append(_empty(idx, word, "Not detected"))
        else:
            r = match_multi(parts, detected, used, max_gap_s=4.0)
            if r:
                start, end, indices, matched, ratio = r
                for i in indices:
                    used.add(i)
                segments.append({
                    "ID":      idx,
                    "Word":    word,
                    "Start_s": start,
                    "End_s":   end,
                    "Note":    f"{matched}/{len(parts)} parts ({ratio}%)",
                    "File":    "—",
                })
            else:
                segments.append(_empty(idx, word, "Not detected"))

    found = sum(1 for s in segments if s["Start_s"] is not None)
    print(f"       -> {found}/{len(segments)} words matched")
    return segments


def _empty(idx, word, note):
    return {"ID": idx, "Word": word,
            "Start_s": None, "End_s": None, "Note": note, "File": "—"}


# ═══════════════════════════════════════════════════════════════════════════
#  STEP 3 – TRIM AUDIO
# ═══════════════════════════════════════════════════════════════════════════
def trim_audio(audio_path, segments, output_dir, padding_ms):
    from pydub import AudioSegment
    audio    = AudioSegment.from_wav(audio_path)
    total_ms = len(audio)
    os.makedirs(output_dir, exist_ok=True)
    saved = 0
    for seg in segments:
        if seg["Start_s"] is None:
            continue
        s_ms  = max(0,        int(seg["Start_s"] * 1000) - padding_ms)
        e_ms  = min(total_ms, int(seg["End_s"]   * 1000) + padding_ms)
        clip  = audio[s_ms:e_ms]
        safe  = re.sub(r"[^\w\-]", "_", seg["Word"]).strip("_")
        fname = f"{seg['ID']:04d}_Word_{safe}.wav"
        fpath = os.path.join(output_dir, fname)
        clip.export(fpath, format="wav")
        seg["File"] = fpath
        saved += 1
    print(f"       -> {saved} clips saved to '{output_dir}/'")


# ═══════════════════════════════════════════════════════════════════════════
#  STEP 4 – EXCEL REPORT
# ═══════════════════════════════════════════════════════════════════════════
def export_excel(segments, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Word Timestamps"

    THIN      = Side(style="thin", color="BBBBBB")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    HDR_FILL  = PatternFill("solid", start_color="1F3864")
    HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    HIT_FILL  = PatternFill("solid", start_color="E2EFDA")   # green  – detected
    MISS_FILL = PatternFill("solid", start_color="FCE4D6")   # orange – not detected
    BODY_FONT = Font(name="Arial", size=11)

    headers    = ["#", "Word",
                  "From (HH:MM:SS.mmm)", "To (HH:MM:SS.mmm)",
                  "From (s)", "To (s)", "Duration (ms)",
                  "Trimmed .wav File", "Status"]
    col_widths = [5, 18, 22, 22, 12, 12, 14, 50, 25]

    # Header row
    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 30

    # Data rows
    for r, seg in enumerate(segments, start=2):
        detected = seg["Start_s"] is not None
        fill     = HIT_FILL if detected else MISS_FILL
        dur      = (round((seg["End_s"] - seg["Start_s"]) * 1000)
                    if detected else None)

        vals = [
            seg["ID"],
            seg["Word"],
            fmt_time(seg["Start_s"]),
            fmt_time(seg["End_s"]),
            seg["Start_s"],
            seg["End_s"],
            dur,
            seg.get("File", "—"),
            seg.get("Note", "—"),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font      = BODY_FONT
            cell.fill      = fill
            cell.border    = BORDER
            cell.alignment = Alignment(vertical="center", horizontal="center"
                                       if c != 2 else "left")
            if c in (5, 6) and isinstance(v, float):
                cell.number_format = "0.000"
            if c == 7 and isinstance(v, int):
                cell.number_format = "#,##0"
        ws.row_dimensions[r].height = 18

    ws.freeze_panes = "A2"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["B"].width = 28
    ws2.column_dimensions["C"].width = 16

    def s2(row, col, value, bold=False, bg=None, fmt=None):
        cell = ws2.cell(row=row, column=col, value=value)
        cell.font = Font(name="Arial", bold=bold, size=11)
        if bg:
            cell.fill = PatternFill("solid", start_color=bg)
        if fmt:
            cell.number_format = fmt
        return cell

    total    = len(segments)
    detected = sum(1 for s in segments if s["Start_s"] is not None)

    s2(1, 2, "Word Detection Summary", bold=True).font = Font(
        name="Arial", bold=True, size=14, color="1F3864")
    s2(3, 2, "Total target words")
    s2(3, 3, total)
    s2(4, 2, "Words detected",     bg="E2EFDA")
    s2(4, 3, detected,             bg="E2EFDA")
    s2(5, 2, "Words NOT detected", bg="FCE4D6")
    s2(5, 3, total - detected,     bg="FCE4D6")
    s2(6, 2, "Detection rate")
    s2(6, 3, "=C5/C4",             fmt="0.0%")

    # List undetected words
    missed = [s["Word"] for s in segments if s["Start_s"] is None]
    if missed:
        s2(8,  2, "Words NOT detected:", bold=True)
        for i, w in enumerate(missed, start=9):
            s2(i, 2, w, bg="FCE4D6")

    wb.save(output_path)
    print(f"\n       Excel saved -> '{output_path}'")


# ═══════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(
        description="Detect target words in a dysarthric .wav file.")
    parser.add_argument("--audio",      required=True,
                        help="Path to input .wav file")
    parser.add_argument("--model",      default="medium",
                        choices=["tiny","base","small","medium",
                                 "large","large-v2","large-v3"],
                        help="Whisper model size (default: medium)")
    parser.add_argument("--method",     default="stable",
                        choices=["stable","whisper"],
                        help="Transcription backend (default: stable)")
    parser.add_argument("--output_dir", default="word_clips",
                        help="Folder to save trimmed word clips (default: word_clips)")
    parser.add_argument("--excel",      default="word_timestamps.xlsx",
                        help="Output Excel filename (default: word_timestamps.xlsx)")
    parser.add_argument("--padding_ms", type=int, default=150,
                        help="Silence padding (ms) around each clip (default: 150)")
    parser.add_argument("--no_trim",    action="store_true",
                        help="Skip audio trimming, only generate Excel")
    parser.add_argument("--save_json",  action="store_true",
                        help="Save all detected words+timestamps as JSON for debugging")
    args = parser.parse_args()

    if not os.path.exists(args.audio):
        print(f"ERROR: File not found – '{args.audio}'")
        sys.exit(1)

    # 1. Transcribe
    if args.method == "stable":
        words = transcribe_stable(args.audio, args.model)
    else:
        words = transcribe_whisper(args.audio, args.model)

    if args.save_json:
        jpath = args.excel.replace(".xlsx", "_all_words.json")
        with open(jpath, "w") as f:
            json.dump(words, f, indent=2)
        print(f"       All detected words -> '{jpath}'")

    # 2. Match target words
    segments = build_segments(words)

    # 3. Trim audio
    if not args.no_trim:
        trim_audio(args.audio, segments, args.output_dir, args.padding_ms)

    # 4. Excel
    export_excel(segments, args.excel)

    found = sum(1 for s in segments if s["Start_s"] is not None)
    print("\n===============================================")
    print(f"  Words found : {found} / {len(segments)}")
    if not args.no_trim:
        print(f"  Clips saved : {args.output_dir}/")
    print(f"  Excel       : {args.excel}")
    print("===============================================\n")


if __name__ == "__main__":
    main()