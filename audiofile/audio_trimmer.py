"""
Dysarthric Speech Audio Trimmer  v4
=====================================
Handles RANDOM ORDER reading — speaker does NOT follow document order.

Strategy:
  - Alphabets/Words  : find the best matching detected word anywhere in audio
  - Sentences        : find a TIME-CLUSTERED group of matching words (close in time)
  - Paragraph        : same cluster approach across the full paragraph

INSTALL:
    pip install openai-whisper stable-ts pydub openpyxl

USAGE:
    python audio_trimmer.py --audio sahil_phase7_sev_english.wav --method stable --model medium
    python audio_trimmer.py --audio file.wav --save_json     # also save raw word list for debug
    python audio_trimmer.py --audio file.wav --no_trim       # Excel only, no audio cutting
"""

import argparse, os, sys, json, re, warnings
warnings.filterwarnings("ignore")


# ═══════════════════════════════════════════════════════════════════════════
#  EXPECTED CONTENT
# ═══════════════════════════════════════════════════════════════════════════
ALPHABETS = list("ABCDEFGHIJKLMNOPQRSTUVWXYZ")

WORDS = [
    "One", "three", "four", "five", "seven", "twelve", "fifteen", "twenty-nine",
    "Their", "If", "Alpha", "Beta", "Delta", "Could", "Adapt", "Circular",
    "Composure", "Footwork", "Journalism", "Python", "Advice", "Choice",
    "Employment", "Immovable", "Massage", "Moisten", "Tree", "Knife",
    "Spoon", "Banana", "Monkey",
]

SENTENCES = [
    ("S1", "Each untimely income loss coincided with the breakdown of a heating system part."),
    ("S2", "Alice's ability to work without supervision is noteworthy."),
    ("S3", "Special task forces rescue hostages from kidnappers."),
    ("S4", "Laugh, dance, and sing if fortune smiles upon you."),
    ("S5", "The same shelter could be built into an embankment or below ground level."),
]

PARAGRAPH = (
    "When the sunlight strikes raindrops in the air they act as a prism and form a rainbow "
    "The rainbow is a division of white light into many beautiful colors "
    "These take the shape of a long round arch with its path high above "
    "and its two ends apparently beyond the horizon "
    "There is according to legend a boiling pot of gold at one end "
    "People look but no one ever finds it "
    "When a man looks for something beyond his reach his friends say he is looking "
    "for the pot of gold at the end of the rainbow "
    "Throughout the centuries people have explained the rainbow in various ways "
    "Some have accepted it as a miracle without physical explanation "
    "To the Hebrews it was a token that there would be no more universal floods "
    "The Greeks used to imagine that it was a sign from the gods to foretell war or heavy rain "
    "The Norsemen considered the rainbow as a bridge over which the gods passed from earth to their home in the sky "
    "Others have tried to explain the phenomenon physically "
    "Aristotle thought that the rainbow was caused by reflection of the suns rays by the rain "
    "Since then physicists have found that it is not reflection but refraction by the raindrops which causes the rainbows "
    "Many complicated ideas about the rainbow have been formed "
    "The difference in the rainbow depends considerably upon the size of the drops "
    "and the width of the colored band increases as the size of the drops increases "
    "The actual primary rainbow observed is said to be the effect of super imposition of a number of bows "
    "If the red of the second bow falls upon the green of the first the result is to give a bow "
    "with an abnormally wide yellow band since red and green light when mixed form yellow "
    "This is a very common type of bow one showing mainly red and yellow with little or no green or blue"
)


# ═══════════════════════════════════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════════════════════════════════
def norm(s):
    return re.sub(r"[^a-z]", "", s.lower())

def norm_words(text):
    return [norm(w) for w in text.split() if norm(w)]

def fmt_time(seconds):
    if seconds is None:
        return "—"
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    s = seconds % 60
    return f"{h:02d}:{m:02d}:{s:06.3f}"

def word_match_score(expected, detected):
    if expected == detected:
        return 1.0
    min_len = min(len(expected), len(detected))
    if min_len >= 4 and expected[:4] == detected[:4]:
        return 0.8
    if len(expected) >= 4 and (expected in detected or detected in expected):
        return 0.6
    if min_len >= 3 and expected[:3] == detected[:3]:
        return 0.5
    return 0.0


# ═══════════════════════════════════════════════════════════════════════════
#  STEP 1 – TRANSCRIBE
# ═══════════════════════════════════════════════════════════════════════════
def transcribe_stable(audio_path, model_size="medium"):
    import stable_whisper
    print(f"\n[1/4] Loading stable-ts '{model_size}' model ...")
    model = stable_whisper.load_model(model_size)
    print(f"[2/4] Transcribing with word-level alignment ...")
    result = model.transcribe(
        audio_path,
        language="en",
        regroup=True,
        suppress_silence=True,
        word_timestamps=True,
        vad=True,
    )
    words = []
    for seg in result.segments:
        for w in seg.words:
            t = norm(w.word)
            if t:
                words.append({"word": t, "start": round(w.start, 3), "end": round(w.end, 3)})
    print(f"       -> {len(words)} words detected")
    return words


def transcribe_whisper(audio_path, model_size="medium"):
    import whisper
    print(f"\n[1/4] Loading Whisper '{model_size}' model ...")
    model = whisper.load_model(model_size)
    print(f"[2/4] Transcribing ...")
    result = model.transcribe(
        audio_path, language="en", word_timestamps=True,
        condition_on_previous_text=False, temperature=0.0,
        no_speech_threshold=0.35,
    )
    words = []
    for seg in result.get("segments", []):
        for w in seg.get("words", []):
            t = norm(w["word"])
            if t:
                words.append({"word": t, "start": round(w["start"], 3), "end": round(w["end"], 3)})
    print(f"       -> {len(words)} words detected")
    return words


# ═══════════════════════════════════════════════════════════════════════════
#  STEP 2 – RANDOM-ORDER MATCHING
# ═══════════════════════════════════════════════════════════════════════════

def match_single_word(expected_norm, detected_words, used_indices):
    best_idx, best_score = None, 0.0
    for i, dw in enumerate(detected_words):
        if i in used_indices:
            continue
        score = word_match_score(expected_norm, dw["word"])
        if score > best_score:
            best_score = score
            best_idx = i
    if best_idx is not None and best_score >= 0.5:
        return best_idx, detected_words[best_idx]["start"], detected_words[best_idx]["end"]
    return None


def match_multi_word(expected_words_norm, detected_words, used_indices,
                     max_gap_s=8.0, min_match_ratio=0.4):
    if not expected_words_norm:
        return None

    # Collect all candidate detected-word positions for any expected word
    candidates = []
    for exp_w in expected_words_norm:
        for i, dw in enumerate(detected_words):
            if i in used_indices:
                continue
            if word_match_score(exp_w, dw["word"]) >= 0.5:
                candidates.append((dw["start"], dw["end"], i))

    if not candidates:
        return None

    candidates.sort(key=lambda x: x[0])

    # Sliding window: find time window with most expected word coverage
    best_window, best_count, best_indices = None, 0, []

    for anchor_idx in range(len(candidates)):
        anchor_t = candidates[anchor_idx][0]
        window = [c for c in candidates if anchor_t <= c[0] <= anchor_t + max_gap_s]
        covered_exp = set()
        for (ts, te, di) in window:
            for exp_w in expected_words_norm:
                if word_match_score(exp_w, detected_words[di]["word"]) >= 0.5:
                    covered_exp.add(exp_w)
        if len(covered_exp) > best_count:
            best_count  = len(covered_exp)
            best_window = window
            best_indices = [c[2] for c in window]

    if best_count < max(1, int(len(expected_words_norm) * min_match_ratio)):
        return None

    start_t = min(c[0] for c in best_window)
    end_t   = max(c[1] for c in best_window)
    ratio   = round(best_count / len(expected_words_norm) * 100)
    return start_t, end_t, best_indices, best_count, ratio


def build_segments(detected_words):
    used = set()
    segments = []
    sid = 1

    print(f"\n[3/4] Matching content to detected words (random order) ...")

    # ALPHABETS
    for letter in ALPHABETS:
        result = match_single_word(norm(letter), detected_words, used)
        if result:
            idx, start, end = result
            used.add(idx)
            segments.append({"Segment_ID": sid, "Type": "Alphabet", "Content": letter,
                              "Start_s": start, "End_s": end, "Note": "", "File": "—"})
        else:
            segments.append(_empty(sid, "Alphabet", letter, "Not detected"))
        sid += 1

    # WORDS
    for word in WORDS:
        exp_norms = norm_words(word)
        if len(exp_norms) == 1:
            result = match_single_word(exp_norms[0], detected_words, used)
            if result:
                idx, start, end = result
                used.add(idx)
                segments.append({"Segment_ID": sid, "Type": "Word", "Content": word,
                                  "Start_s": start, "End_s": end, "Note": "", "File": "—"})
            else:
                segments.append(_empty(sid, "Word", word, "Not detected"))
        else:
            result = match_multi_word(exp_norms, detected_words, used, max_gap_s=5.0)
            if result:
                start, end, indices, matched, ratio = result
                for i in indices: used.add(i)
                segments.append({"Segment_ID": sid, "Type": "Word", "Content": word,
                                  "Start_s": start, "End_s": end,
                                  "Note": f"{matched}/{len(exp_norms)} parts ({ratio}%)", "File": "—"})
            else:
                segments.append(_empty(sid, "Word", word, "Not detected"))
        sid += 1

    # SENTENCES
    for label, sentence in SENTENCES:
        exp_norms = norm_words(sentence)
        result = match_multi_word(exp_norms, detected_words, used,
                                  max_gap_s=30.0, min_match_ratio=0.35)
        if result:
            start, end, indices, matched, ratio = result
            for i in indices: used.add(i)
            segments.append({"Segment_ID": sid, "Type": f"Sentence {label}", "Content": sentence,
                              "Start_s": start, "End_s": end,
                              "Note": f"{matched}/{len(exp_norms)} words ({ratio}%)", "File": "—"})
        else:
            segments.append(_empty(sid, f"Sentence {label}", sentence, "Not detected"))
        sid += 1

    # PARAGRAPH
    exp_norms = norm_words(PARAGRAPH)
    result = match_multi_word(exp_norms, detected_words, used,
                              max_gap_s=300.0, min_match_ratio=0.25)
    if result:
        start, end, indices, matched, ratio = result
        for i in indices: used.add(i)
        segments.append({"Segment_ID": sid, "Type": "Paragraph",
                          "Content": PARAGRAPH[:120] + "...",
                          "Start_s": start, "End_s": end,
                          "Note": f"{matched}/{len(exp_norms)} words ({ratio}%)", "File": "—"})
    else:
        segments.append(_empty(sid, "Paragraph", PARAGRAPH[:120] + "...", "Not detected"))

    detected_count = sum(1 for s in segments if s["Start_s"] is not None)
    print(f"       -> {detected_count}/{len(segments)} segments matched")
    return segments


def _empty(sid, stype, content, note):
    return {"Segment_ID": sid, "Type": stype, "Content": content,
            "Start_s": None, "End_s": None, "Note": note, "File": "—"}


# ═══════════════════════════════════════════════════════════════════════════
#  STEP 3 – TRIM AUDIO
# ═══════════════════════════════════════════════════════════════════════════
def trim_audio(audio_path, segments, output_dir, padding_ms=200):
    from pydub import AudioSegment
    audio    = AudioSegment.from_wav(audio_path)
    total_ms = len(audio)
    os.makedirs(output_dir, exist_ok=True)
    saved = 0
    for seg in segments:
        if seg["Start_s"] is None:
            continue
        s_ms = max(0,        int(seg["Start_s"] * 1000) - padding_ms)
        e_ms = min(total_ms, int(seg["End_s"]   * 1000) + padding_ms)
        clip = audio[s_ms:e_ms]
        safe  = re.sub(r"[^\w\-]", "_", seg["Content"][:35]).strip("_")
        fname = f"{seg['Segment_ID']:04d}_{seg['Type'].replace(' ','_')}_{safe}.wav"
        fpath = os.path.join(output_dir, fname)
        clip.export(fpath, format="wav")
        seg["File"] = fpath
        saved += 1
    print(f"       -> {saved} clips saved to '{output_dir}/'")


# ═══════════════════════════════════════════════════════════════════════════
#  STEP 4 – EXCEL REPORT
# ═══════════════════════════════════════════════════════════════════════════
def export_excel(segments, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Audio Segments"

    THIN   = Side(style="thin", color="BBBBBB")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    TYPE_COLOR = {
        "Alphabet":  "FFF2CC",
        "Word":      "E2EFDA",
        "Sentence":  "DDEBF7",
        "Paragraph": "FCE4D6",
    }

    headers    = ["ID", "Type", "Content / Text Spoken",
                  "From (HH:MM:SS.mmm)", "To (HH:MM:SS.mmm)",
                  "From (s)", "To (s)", "Duration (ms)",
                  "Trimmed .wav File", "Match Info"]
    col_widths = [5, 15, 62, 22, 22, 11, 11, 14, 55, 28]

    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", start_color="1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 32

    for r, seg in enumerate(segments, start=2):
        tk   = seg["Type"].split()[0]
        fill = PatternFill("solid", start_color=TYPE_COLOR.get(tk, "FFFFFF"))
        dur  = (round((seg["End_s"] - seg["Start_s"]) * 1000)
                if seg["Start_s"] is not None else None)
        vals = [
            seg["Segment_ID"], seg["Type"], seg["Content"],
            fmt_time(seg["Start_s"]), fmt_time(seg["End_s"]),
            seg["Start_s"], seg["End_s"], dur,
            seg.get("File", "—"), seg.get("Note", ""),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = fill
            cell.border    = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(c in (3, 9)))
            if c in (6, 7) and isinstance(v, float):
                cell.number_format = "0.000"
            if c == 8 and isinstance(v, int):
                cell.number_format = "#,##0"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["B"].width = 34
    ws2.column_dimensions["C"].width = 18

    def s2(row, col, value, bold=False, bg=None, fmt=None):
        cell = ws2.cell(row=row, column=col, value=value)
        cell.font = Font(name="Arial", bold=bold, size=11)
        if bg:
            cell.fill = PatternFill("solid", start_color=bg)
        if fmt:
            cell.number_format = fmt
        return cell

    total    = len(segments)
    detected = sum(1 for s in segments if s["Start_s"] is not None)

    s2(1, 2, "Dysarthric Speech Segmentation Report", bold=True).font = Font(
        name="Arial", bold=True, size=14, color="1F3864")
    s2(3, 2, "Total segments expected")
    s2(3, 3, total)
    s2(4, 2, "Segments detected",  bg="E2EFDA")
    s2(4, 3, detected,             bg="E2EFDA")
    s2(5, 2, "Segments NOT found", bg="FCE4D6")
    s2(5, 3, total - detected,     bg="FCE4D6")
    s2(6, 2, "Detection rate")
    s2(6, 3, "=C5/C4",             fmt="0.0%")
    s2(8, 2, "BREAKDOWN BY TYPE",  bold=True, bg="1F3864").font = Font(
        name="Arial", bold=True, color="FFFFFF", size=11)
    s2(8, 3, "Detected / Total",   bold=True, bg="1F3864").font = Font(
        name="Arial", bold=True, color="FFFFFF", size=11)

    type_data = {}
    for seg in segments:
        tk = seg["Type"].split()[0]
        type_data.setdefault(tk, [0, 0])
        type_data[tk][1] += 1
        if seg["Start_s"] is not None:
            type_data[tk][0] += 1

    for i, (tk, (det, tot)) in enumerate(type_data.items(), start=9):
        color = TYPE_COLOR.get(tk, "FFFFFF")
        s2(i, 2, tk,               bg=color)
        s2(i, 3, f"{det} / {tot}", bg=color)

    wb.save(output_path)
    print(f"\n       Excel saved -> '{output_path}'")


# ═══════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--audio",      required=True)
    parser.add_argument("--model",      default="medium",
                        choices=["tiny","base","small","medium","large","large-v2","large-v3"])
    parser.add_argument("--method",     default="stable", choices=["stable","whisper"])
    parser.add_argument("--output_dir", default="trimmed_segments")
    parser.add_argument("--excel",      default="audio_timestamps.xlsx")
    parser.add_argument("--padding_ms", type=int, default=200)
    parser.add_argument("--no_trim",    action="store_true")
    parser.add_argument("--save_json",  action="store_true")
    args = parser.parse_args()

    if not os.path.exists(args.audio):
        print(f"ERROR: File not found - '{args.audio}'"); sys.exit(1)

    if args.method == "stable":
        words = transcribe_stable(args.audio, args.model)
    else:
        words = transcribe_whisper(args.audio, args.model)

    if args.save_json:
        jpath = args.excel.replace(".xlsx", "_words.json")
        with open(jpath, "w") as f:
            json.dump(words, f, indent=2)
        print(f"       Raw word list -> '{jpath}'")

    segments = build_segments(words)

    if not args.no_trim:
        trim_audio(args.audio, segments, args.output_dir, args.padding_ms)

    export_excel(segments, args.excel)

    detected = sum(1 for s in segments if s["Start_s"] is not None)
    print("\n===============================================")
    print(f"  Segments : {detected}/{len(segments)} detected")
    if not args.no_trim:
        print(f"  Clips    : {args.output_dir}/")
    print(f"  Excel    : {args.excel}")
    print("===============================================\n")


if __name__ == "__main__":
    main()