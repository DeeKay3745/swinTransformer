"""
Marathi Dysarthric Speech Trimmer
===================================
Detects Marathi varnamala, words, sentences and paragraph
from a .wav file using Whisper and exports trimmed clips + Excel.

LANGUAGE : Marathi (Devanagari script)
MODEL    : Whisper – supports Marathi natively (language="mr")

INSTALL:
    pip install openai-whisper stable-ts pydub openpyxl

USAGE:
    python marathi_trimmer.py --audio your_file.wav --method stable --model medium
    python marathi_trimmer.py --audio your_file.wav --save_json
    python marathi_trimmer.py --audio your_file.wav --no_trim
"""

import argparse, os, sys, json, re, unicodedata, warnings
from difflib import SequenceMatcher
warnings.filterwarnings("ignore")


# ═══════════════════════════════════════════════════════════════════════════
#  CONTENT  (from Abhishek_Marathi.docx – exact order)
# ═══════════════════════════════════════════════════════════════════════════

# ── Vowels (स्वर) ──────────────────────────────────────────────────────────
VOWELS = ["अ","आ","ॲ","ऑ","इ","ई","उ","ऊ","ऋ","ऌ","ए","ऐ","ओ","औ","अं","अः"]

# ── Consonants (व्यंजन) ───────────────────────────────────────────────────
CONSONANTS = [
    "क्","ख्","ग्","घ्","ङ्",
    "च्","छ्","ज्","झ्","ञ्",
    "ट्","ठ्","ड्","ढ्","ण्",
    "त्","थ्","द्","ध्","न्",
    "प","फ्","ब्","भ्","म्",
    "य्","र्","ल्","व्","श्","ष्","स्","ह्","ळ्",
]

VARNAMALA = VOWELS + CONSONANTS   # 16 vowels + 34 consonants = 50 characters

# ── Words (शब्द) – 50 words ────────────────────────────────────────────────
WORDS = [
    "एक","तीन","चार","पाच","सात","बारा","पंधरा","एकोणतीस",
    "आई","वडील","शाळा","आंबा","घर",
    "बालपण","भाऊ","नवरा","बहीण","गणपती",
    "पुस्तक","शिक्षक","टिपणवही","डोळे","गुडघा",
    "दसरा","उत्सव","उन्हाळा","पावसाळी",
    "हिवाळा","बोट","चित्रपट","राग","शेती",
    "चहा","किल्ला","कान","फळा","खडू",
    "ज्ञानेश्वर","कपाळ","पिके","ऊस","चटणी",
    "हळद","दगड","भांडे","लोणचे","पोट",
    "गळा","साडी","सासर","निबंध","शिकवणी",
]

# ── Sentences (वाक्ये) ────────────────────────────────────────────────────
SENTENCES = [
    ("S1", "कंटाळवाणी कादंबरी म्हणजे झोपेची उत्तम गोळी."),
    ("S2", "जेफला वाटले की आपण सेंट्रीफ्यूज खरेदीच्या बाजूने युक्तिवाद केला."),
    ("S3", "ती नैसर्गिकरित्या तिच्या शारीरिक स्वरूपाशिवाय इतर सर्व गोष्टींबद्दल ढिसाळ होती का?"),
    ("S4", "घरातील निवांत वातावरण तिला या गुणांवर मात करण्यास मदत करण्यासाठी पुरेसे आहे का?"),
    ("S5", "हाच निवारा बंधाऱ्यात किंवा जमिनीच्या खाली बांधता येऊ शकतो."),
]

# ── Paragraph sentences (परिच्छेद) – split for accurate localisation ──────
PARA_SENTENCES = [
    "शिवाजी भोसले यांचा जन्म १९ फेब्रुवारी १६३० रोजी शहाजी भोसले आणि जिजाबाई यांच्या पोटी पुणे जिल्ह्यातील जुन्नर शहराजवळील शिवनेरी किल्ल्यावर झाला.",
    "शिवाजीचे वडील शहाजी हे विजापुरी सल्तनतच्या सेवेत होते सेनापती म्हणून विजापूर अहमदनगर आणि गोलकोंडा यांच्यातील त्रिपक्षीय संघटना.",
    "पुण्याजवळ त्यांची जयगिरदारीही होती.",
    "शिवरायांच्या आई जिजाबाई या सिंदखेडचे नेते लखुजीराव जाधव यांच्या कन्या आणि अत्यंत धार्मिक स्त्री होत्या.",
    "शिवाजी लहानपणापासूनच जन्मजात नेता होता.",
    "त्याने शिवनेरी किल्ल्यांभोवती असलेल्या सह्याद्री पर्वताचे अन्वेषण केले आणि त्याच्या हाताच्या मागील भागासारखे क्षेत्र ओळखले.",
    "तो १५ वर्षांचा होता तोपर्यंत त्याने मावळ प्रदेशातील विश्वासू सैनिकांचा एक गट जमा केला होता ज्यांनी नंतर त्याच्या सुरुवातीच्या विजयांमध्ये मदत केली.",
    "शिवाजीचा विजापुरी सल्तनतशी झालेला संघर्ष आणि त्याच्या सततच्या विजयामुळे तो मुघल सम्राट औरंगजेबाच्या रडारखाली आला.",
    "औरंगजेबाने त्याला आपल्या साम्राज्यवादी हेतूच्या विस्ताराचा धोका म्हणून पाहिले आणि मराठ्यांचा धोका नष्ट करण्यासाठी आपले प्रयत्न केंद्रित केले.",
    "शिवाजीने आपल्या दरबारात फारसी या सध्याच्या शाही भाषेऐवजी मराठी आणि संस्कृतच्या वापराचा जोरदार प्रचार केला.",
    "शिवाजी हे स्वत: धर्माभिमानी असले तरी त्यांनी त्यांच्या राजवटीत सर्व धर्मांसाठी सहिष्णुतेचा पुरस्कार केला.",
    "ते जातिभेदाच्या विरोधात होते आणि त्यांच्या दरबारात सर्व जातीतील लोकांना कामावर ठेवत.",
    "शेतकरी आणि राज्य यांच्यातील मध्यस्थांची गरज दूर करणारी रयतवारी प्रणाली त्यांनी सुरू केली.",
]

PARAGRAPH_FULL = " ".join(PARA_SENTENCES)


# ═══════════════════════════════════════════════════════════════════════════
#  DEVANAGARI HELPERS
#  Standard Latin prefix tricks don't apply to Devanagari.
#  We use:
#    1. Unicode NFC normalisation (handles half-letters, matras, nuktas)
#    2. SequenceMatcher ratio for fuzzy character overlap
#    3. Strip punctuation (। , ? ! . " ')
# ═══════════════════════════════════════════════════════════════════════════
def nfc(s):
    """Unicode NFC normalise + strip whitespace."""
    return unicodedata.normalize("NFC", s.strip())

def clean(s):
    """NFC + remove all punctuation/digits, keep only Devanagari chars."""
    s = nfc(s)
    # keep Devanagari Unicode block U+0900–U+097F + ZWJ/ZWNJ
    s = re.sub(r"[^\u0900-\u097F\u200C\u200D]", "", s)
    return s

def clean_words(text):
    """Split text into cleaned Devanagari tokens."""
    text = nfc(text.replace("\xa0", " "))
    tokens = re.split(r"[\s।,\.\?\!\"\'\(\)]+", text)
    return [clean(t) for t in tokens if clean(t)]

def deva_similarity(a, b):
    """
    Character-level similarity score 0..1 between two Devanagari words.
    Uses SequenceMatcher on the cleaned strings.
    Exact match → 1.0, totally different → 0.0
    """
    a, b = clean(a), clean(b)
    if not a or not b:
        return 0.0
    if a == b:
        return 1.0
    # SequenceMatcher ratio = 2*M / T  (M=matching chars, T=total chars)
    ratio = SequenceMatcher(None, a, b).ratio()
    return ratio

def fmt_time(seconds):
    if seconds is None:
        return "—"
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    s = seconds % 60
    return f"{h:02d}:{m:02d}:{s:06.3f}"


# ═══════════════════════════════════════════════════════════════════════════
#  STEP 1 – TRANSCRIBE  (language="mr" for Marathi)
# ═══════════════════════════════════════════════════════════════════════════
def transcribe_stable(audio_path, model_size):
    import stable_whisper
    print(f"\n[1/4] Loading stable-ts '{model_size}' model ...")
    model = stable_whisper.load_model(model_size)
    print(f"[2/4] Transcribing Marathi audio ...")
    result = model.transcribe(
        audio_path,
        language="mr",          # ← Marathi
        regroup=True,
        suppress_silence=True,
        word_timestamps=True,
        vad=True,
    )
    words = []
    for seg in result.segments:
        for w in seg.words:
            t = clean(w.word)
            if t:
                words.append({
                    "word":  t,
                    "raw":   nfc(w.word),
                    "start": round(w.start, 3),
                    "end":   round(w.end,   3),
                })
    print(f"       -> {len(words)} Marathi words detected")
    return words


def transcribe_whisper(audio_path, model_size):
    import whisper
    print(f"\n[1/4] Loading Whisper '{model_size}' model ...")
    model = whisper.load_model(model_size)
    print(f"[2/4] Transcribing Marathi audio ...")
    result = model.transcribe(
        audio_path,
        language="mr",          # ← Marathi
        word_timestamps=True,
        condition_on_previous_text=False,
        temperature=0.0,
        no_speech_threshold=0.35,
    )
    words = []
    for seg in result.get("segments", []):
        for w in seg.get("words", []):
            t = clean(w["word"])
            if t:
                words.append({
                    "word":  t,
                    "raw":   nfc(w["word"]),
                    "start": round(w["start"], 3),
                    "end":   round(w["end"],   3),
                })
    print(f"       -> {len(words)} Marathi words detected")
    return words


# ═══════════════════════════════════════════════════════════════════════════
#  STEP 2A – MATCH SINGLE TOKEN
# ═══════════════════════════════════════════════════════════════════════════
MATCH_THRESHOLD = 0.60   # minimum similarity to accept a match

def match_single(expected, detected, used):
    """
    Find the best unused detected word for the expected Devanagari token.
    Returns (index, start, end, score) or None.
    """
    exp_clean = clean(expected)
    best_idx, best_score = None, 0.0
    for i, dw in enumerate(detected):
        if i in used:
            continue
        score = deva_similarity(exp_clean, dw["word"])
        if score > best_score:
            best_score, best_idx = score, i
    if best_idx is not None and best_score >= MATCH_THRESHOLD:
        d = detected[best_idx]
        return best_idx, d["start"], d["end"], round(best_score * 100)
    return None


# ═══════════════════════════════════════════════════════════════════════════
#  STEP 2B – MATCH MULTI-WORD (sentences / paragraph)
#  Anchor on long/distinctive words, then expand window
# ═══════════════════════════════════════════════════════════════════════════

# Common Marathi stop words (too frequent to use as anchors)
STOP_MR = {
    clean(w) for w in [
        "आणि","की","हे","ते","ती","तो","या","यांचा","यांच्या","आहे","होते",
        "होता","होती","केले","केली","केला","म्हणून","त्याचे","त्याची","त्याच्या",
        "आपल्या","त्यांनी","त्यांच्या","त्यांची","एक","अनेक","सर्व","इतर",
        "असे","असा","असलेल्या","असले","नंतर","तेव्हा","जेव्हा","जिथे",
        "परंतु","तथापि","मात्र","पण","जे","जो","जी","ज्यांनी","ज्याने",
    ]
}

def distinctive_mr(text, min_chars=4):
    """Return cleaned tokens that are long and not stop words."""
    return [w for w in clean_words(text)
            if len(w) >= min_chars and w not in STOP_MR]

def match_multi(sentence_text, detected, used,
                max_window_s=60.0, min_ratio=0.25):
    """
    Match a multi-word sentence/paragraph section.
    1. Find anchor positions using DISTINCTIVE words.
    2. For each anchor, count ALL sentence words covered in window.
    3. Return the best window.
    """
    all_exp   = clean_words(sentence_text)
    key_exp   = distinctive_mr(sentence_text)

    if not all_exp:
        return None

    # ── collect candidate anchors from distinctive words ──
    anchors = []
    for exp_w in (key_exp if key_exp else all_exp):
        for i, dw in enumerate(detected):
            if i in used:
                continue
            if deva_similarity(exp_w, dw["word"]) >= MATCH_THRESHOLD:
                anchors.append((dw["start"], i))

    if not anchors:
        return None

    # Deduplicate anchors per second
    seen, deduped = set(), []
    for (t, i) in sorted(anchors):
        bucket = int(t)
        if bucket not in seen:
            seen.add(bucket)
            deduped.append((t, i))

    best_result, best_score = None, 0

    for anchor_t, _ in deduped:
        # All unused detected words in window [anchor - 2s … anchor + max_window_s]
        window = [(i, dw["word"], dw["start"], dw["end"])
                  for i, dw in enumerate(detected)
                  if i not in used
                  and anchor_t - 2.0 <= dw["start"] <= anchor_t + max_window_s]

        if not window:
            continue

        # Greedy: match each expected word to best window word (no reuse)
        used_in_w = set()
        matched_words = []
        for exp_w in all_exp:
            best_i, best_s = None, 0.0
            for (wi, wword, wstart, wend) in window:
                if wi in used_in_w:
                    continue
                s = deva_similarity(exp_w, wword)
                if s > best_s:
                    best_s, best_i = s, wi
            if best_i is not None and best_s >= MATCH_THRESHOLD:
                used_in_w.add(best_i)
                matched_words.append(best_i)

        score = len(matched_words) / len(all_exp)
        if score > best_score:
            best_score = score
            matched_window = [w for w in window if w[0] in matched_words]
            if matched_window:
                start_t = min(w[2] for w in matched_window)
                end_t   = max(w[3] for w in matched_window)
                best_result = (start_t, end_t, matched_words,
                               len(matched_words), len(all_exp),
                               round(score * 100))

    if best_result is None or best_result[5] < min_ratio * 100:
        return None
    return best_result


def match_paragraph(detected, used):
    """Match paragraph sentence by sentence then span the full range."""
    all_starts, all_ends, all_indices = [], [], []
    total_matched, total_words = 0, len(clean_words(PARAGRAPH_FULL))
    matched_sents = 0

    for sent in PARA_SENTENCES:
        r = match_multi(sent, detected, used,
                        max_window_s=90.0, min_ratio=0.20)
        if r:
            s, e, idxs, matched, total, ratio = r
            all_starts.append(s)
            all_ends.append(e)
            all_indices.extend(idxs)
            total_matched += matched
            matched_sents += 1
            for i in idxs:
                used.add(i)

    if not all_starts:
        return None

    ratio = round(total_matched / total_words * 100) if total_words else 0
    note  = (f"{matched_sents}/{len(PARA_SENTENCES)} sentences, "
             f"{total_matched}/{total_words} words ({ratio}%)")
    return min(all_starts), max(all_ends), list(set(all_indices)), note


# ═══════════════════════════════════════════════════════════════════════════
#  STEP 2 – BUILD ALL SEGMENTS
# ═══════════════════════════════════════════════════════════════════════════
def build_segments(detected):
    used = set()
    segments = []
    sid = 1

    print(f"\n[3/4] Matching segments ...")

    # ── VARNAMALA (vowels then consonants) ────────────────────────────────
    for char in VARNAMALA:
        r = match_single(char, detected, used)
        if r:
            idx, start, end, score = r
            used.add(idx)
            label = "Vowel" if char in VOWELS else "Consonant"
            segments.append(_seg(sid, label, char, start, end,
                                 f"score {score}%"))
        else:
            label = "Vowel" if char in VOWELS else "Consonant"
            segments.append(_empty(sid, label, char, "Not detected"))
        sid += 1

    # ── WORDS ─────────────────────────────────────────────────────────────
    for word in WORDS:
        parts = clean_words(word)
        if len(parts) <= 1:
            r = match_single(word, detected, used)
            if r:
                idx, start, end, score = r
                used.add(idx)
                segments.append(_seg(sid, "Word", word, start, end,
                                     f"score {score}%"))
            else:
                segments.append(_empty(sid, "Word", word, "Not detected"))
        else:
            # compound word — use cluster match
            r = match_multi(word, detected, used,
                            max_window_s=5.0, min_ratio=0.5)
            if r:
                start, end, idxs, matched, total, ratio = r
                for i in idxs: used.add(i)
                segments.append(_seg(sid, "Word", word, start, end,
                                     f"{matched}/{total} parts ({ratio}%)"))
            else:
                segments.append(_empty(sid, "Word", word, "Not detected"))
        sid += 1

    # ── SENTENCES ─────────────────────────────────────────────────────────
    for label, sentence in SENTENCES:
        r = match_multi(sentence, detected, used,
                        max_window_s=60.0, min_ratio=0.25)
        if r:
            start, end, idxs, matched, total, ratio = r
            for i in idxs: used.add(i)
            segments.append(_seg(sid, f"Sentence {label}", sentence,
                                 start, end,
                                 f"{matched}/{total} words ({ratio}%)"))
        else:
            segments.append(_empty(sid, f"Sentence {label}",
                                   sentence, "Not detected"))
        sid += 1

    # ── PARAGRAPH ─────────────────────────────────────────────────────────
    r = match_paragraph(detected, used)
    if r:
        start, end, idxs, note = r
        segments.append(_seg(sid, "Paragraph",
                             PARAGRAPH_FULL[:100] + "...",
                             start, end, note))
    else:
        segments.append(_empty(sid, "Paragraph",
                               PARAGRAPH_FULL[:100] + "...", "Not detected"))

    found = sum(1 for s in segments if s["Start_s"] is not None)
    print(f"       -> {found}/{len(segments)} segments matched")
    return segments


def _seg(sid, stype, content, s, e, note):
    return {"Segment_ID": sid, "Type": stype, "Content": content,
            "Start_s": s, "End_s": e, "Note": note, "File": "—"}

def _empty(sid, stype, content, note):
    return {"Segment_ID": sid, "Type": stype, "Content": content,
            "Start_s": None, "End_s": None, "Note": note, "File": "—"}


# ═══════════════════════════════════════════════════════════════════════════
#  STEP 3 – TRIM AUDIO
# ═══════════════════════════════════════════════════════════════════════════
def trim_audio(audio_path, segments, output_dir, padding_ms):
    from pydub import AudioSegment
    audio    = AudioSegment.from_wav(audio_path)
    total_ms = len(audio)
    os.makedirs(output_dir, exist_ok=True)
    saved = 0
    for seg in segments:
        if seg["Start_s"] is None:
            continue
        s_ms  = max(0,        int(seg["Start_s"] * 1000) - padding_ms)
        e_ms  = min(total_ms, int(seg["End_s"]   * 1000) + padding_ms)
        clip  = audio[s_ms:e_ms]
        # Use segment ID + romanised type for safe filename
        safe  = re.sub(r"[^\w\-]", "_", seg["Type"].replace(" ", "_"))
        fname = f"{seg['Segment_ID']:04d}_{safe}.wav"
        fpath = os.path.join(output_dir, fname)
        clip.export(fpath, format="wav")
        seg["File"] = fpath
        saved += 1
    print(f"       -> {saved} clips saved to '{output_dir}/'")


# ═══════════════════════════════════════════════════════════════════════════
#  STEP 4 – EXCEL REPORT
# ═══════════════════════════════════════════════════════════════════════════
def export_excel(segments, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Audio Segments"

    THIN   = Side(style="thin", color="BBBBBB")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    TYPE_COLOR = {
        "Vowel":     "E8F4FD",
        "Consonant": "FFF9E6",
        "Word":      "E2EFDA",
        "Sentence":  "DDEBF7",
        "Paragraph": "FCE4D6",
    }

    headers    = ["ID", "Type", "Content (Marathi)",
                  "From (HH:MM:SS.mmm)", "To (HH:MM:SS.mmm)",
                  "From (s)", "To (s)", "Duration (ms)",
                  "Trimmed .wav File", "Match Info"]
    col_widths = [5, 14, 45, 22, 22, 11, 11, 14, 45, 28]

    for c, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", start_color="1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 32

    for r, seg in enumerate(segments, start=2):
        tk   = seg["Type"].split()[0]
        fill = PatternFill("solid", start_color=TYPE_COLOR.get(tk, "FFFFFF"))
        dur  = (round((seg["End_s"] - seg["Start_s"]) * 1000)
                if seg["Start_s"] is not None else None)
        vals = [
            seg["Segment_ID"], seg["Type"], seg["Content"],
            fmt_time(seg["Start_s"]), fmt_time(seg["End_s"]),
            seg["Start_s"], seg["End_s"], dur,
            seg.get("File", "—"), seg.get("Note", ""),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font      = Font(name="Nirmala UI", size=11) if c == 3 \
                             else Font(name="Arial",    size=10)
            cell.fill      = fill
            cell.border    = BORDER
            cell.alignment = Alignment(vertical="center",
                                       wrap_text=(c in (3, 9)))
            if c in (6, 7) and isinstance(v, float):
                cell.number_format = "0.000"
            if c == 8 and isinstance(v, int):
                cell.number_format = "#,##0"
        ws.row_dimensions[r].height = 20

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ── Summary ──────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["B"].width = 34
    ws2.column_dimensions["C"].width = 18

    def s2(row, col, value, bold=False, bg=None, fmt=None):
        cell = ws2.cell(row=row, column=col, value=value)
        cell.font = Font(name="Arial", bold=bold, size=11)
        if bg:
            cell.fill = PatternFill("solid", start_color=bg)
        if fmt:
            cell.number_format = fmt
        return cell

    total    = len(segments)
    detected = sum(1 for s in segments if s["Start_s"] is not None)

    s2(1, 2, "Marathi Speech Segmentation Report", bold=True).font = Font(
        name="Arial", bold=True, size=14, color="1F3864")
    s2(3, 2, "Total segments expected")
    s2(3, 3, total)
    s2(4, 2, "Segments detected",  bg="E2EFDA")
    s2(4, 3, detected,             bg="E2EFDA")
    s2(5, 2, "Segments NOT found", bg="FCE4D6")
    s2(5, 3, total - detected,     bg="FCE4D6")
    s2(6, 2, "Detection rate")
    s2(6, 3, "=C5/C4",             fmt="0.0%")
    s2(8, 2, "BREAKDOWN BY TYPE",  bold=True, bg="1F3864").font = Font(
        name="Arial", bold=True, color="FFFFFF", size=11)
    s2(8, 3, "Detected / Total",   bold=True, bg="1F3864").font = Font(
        name="Arial", bold=True, color="FFFFFF", size=11)

    type_data = {}
    for seg in segments:
        tk = seg["Type"].split()[0]
        type_data.setdefault(tk, [0, 0])
        type_data[tk][1] += 1
        if seg["Start_s"] is not None:
            type_data[tk][0] += 1

    for i, (tk, (det, tot)) in enumerate(type_data.items(), start=9):
        color = TYPE_COLOR.get(tk, "FFFFFF")
        s2(i, 2, tk,               bg=color)
        s2(i, 3, f"{det} / {tot}", bg=color)

    # List undetected
    missed = [s["Content"] for s in segments if s["Start_s"] is None]
    if missed:
        row = 9 + len(type_data) + 2
        s2(row, 2, "Not Detected:", bold=True)
        for j, m in enumerate(missed, start=row+1):
            cell = ws2.cell(row=j, column=2, value=m)
            cell.font = Font(name="Nirmala UI", size=11)
            cell.fill = PatternFill("solid", start_color="FCE4D6")

    wb.save(output_path)
    print(f"\n       Excel saved -> '{output_path}'")


# ═══════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(
        description="Marathi dysarthric speech trimmer with word-level timestamps.")
    parser.add_argument("--audio",      required=True,
                        help="Path to input .wav file")
    parser.add_argument("--model",      default="medium",
                        choices=["tiny","base","small","medium",
                                 "large","large-v2","large-v3"],
                        help="Whisper model size (large-v2 recommended for Marathi)")
    parser.add_argument("--method",     default="stable",
                        choices=["stable","whisper"])
    parser.add_argument("--output_dir", default="marathi_clips")
    parser.add_argument("--excel",      default="marathi_timestamps.xlsx")
    parser.add_argument("--padding_ms", type=int, default=200)
    parser.add_argument("--no_trim",    action="store_true")
    parser.add_argument("--save_json",  action="store_true",
                        help="Save all detected words as JSON for debugging")
    args = parser.parse_args()

    if not os.path.exists(args.audio):
        print(f"ERROR: File not found - '{args.audio}'")
        sys.exit(1)

    # 1. Transcribe
    if args.method == "stable":
        words = transcribe_stable(args.audio, args.model)
    else:
        words = transcribe_whisper(args.audio, args.model)

    if args.save_json:
        jpath = args.excel.replace(".xlsx", "_words.json")
        with open(jpath, "w", encoding="utf-8") as f:
            json.dump(words, f, ensure_ascii=False, indent=2)
        print(f"       Raw word list -> '{jpath}'")

    # 2. Match
    segments = build_segments(words)

    # 3. Trim
    if not args.no_trim:
        trim_audio(args.audio, segments, args.output_dir, args.padding_ms)

    # 4. Excel
    export_excel(segments, args.excel)

    found = sum(1 for s in segments if s["Start_s"] is not None)
    print("\n===============================================")
    print(f"  Language : Marathi (mr)")
    print(f"  Segments : {found}/{len(segments)} detected")
    if not args.no_trim:
        print(f"  Clips    : {args.output_dir}/")
    print(f"  Excel    : {args.excel}")
    print("===============================================\n")


if __name__ == "__main__":
    main()