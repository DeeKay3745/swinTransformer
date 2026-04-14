import argparse
import os
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from pydub import AudioSegment

#format for the column is hh:mm:ss.000
def parse_time(ts: str) -> float:
    ts = ts.strip()
    m = re.match(r'(\d+):(\d+):(\d+(?:\.\d+)?)', ts)
    if not m:
        raise ValueError(f"Cannot parse timestamp: '{ts}'")
    h, mi, s = int(m.group(1)), int(m.group(2)), float(m.group(3))
    return (h * 3600 + mi * 60 + s) * 1000  # milliseconds


def safe_filename(text: str, max_len: int = 50) -> str:
    text = text.strip()
    text = re.sub(r'[^\w\s-]', '', text)   
    text = re.sub(r'\s+', '_', text)        
    return text[:max_len]


def read_excel(xlsx_path: str, type_filter: str = "all") -> list:
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if row[0] is None:
            continue
        seg_type = str(row[0]).strip()
        content = str(row[1]).strip() if row[1] else ""
        start_ts = str(row[2]).strip() if row[2] else ""
        end_ts = str(row[3]).strip() if row[3] else ""

        if not start_ts or not end_ts:
            continue

        if type_filter != "all" and seg_type.lower() != type_filter.lower():
            continue

        rows.append({
            'index': i,
            'type': seg_type,
            'content': content,
            'start_ts': start_ts,
            'end_ts': end_ts,
        })

    wb.close()
    return rows


def split_audio(wav_path: str, segments: list, output_dir: str, out_format: str = "wav"):
    """Split the WAV file according to segments and save to output_dir."""
    print(f"Loading audio: {wav_path}")
    audio = AudioSegment.from_wav(wav_path)
    total_ms = len(audio)
    print(f"  Duration: {total_ms / 1000:.2f}s")

    os.makedirs(output_dir, exist_ok=True)

    # Create sub-folders per type
    types_seen = set(seg['type'] for seg in segments)
    for t in types_seen:
        os.makedirs(os.path.join(output_dir, t), exist_ok=True)

    count = 0
    for seg in segments:
        start_ms = parse_time(seg['start_ts'])
        end_ms = parse_time(seg['end_ts'])

   
        start_ms = max(0, min(start_ms, total_ms))
        end_ms = max(start_ms, min(end_ms, total_ms))

        if end_ms - start_ms < 10:  
            print(f"  SKIP (too short): {seg['type']} — {seg['content'][:40]}")
            continue

        chunk = audio[start_ms:end_ms]

     
        name = safe_filename(seg['content'])
        filename = f"{count + 1:04d}_{seg['type']}_{name}.{out_format}"
        filepath = os.path.join(output_dir, seg['type'], filename)

        chunk.export(filepath, format=out_format)
        dur = (end_ms - start_ms) / 1000
        print(f"  [{count + 1:4d}] {seg['type']:12s} | {dur:6.2f}s | {filename}")
        count += 1

    print(f"\nDone! {count} segments saved to: {output_dir}")
    for t in sorted(types_seen):
        folder = os.path.join(output_dir, t)
        n = len([f for f in os.listdir(folder) if f.endswith(f'.{out_format}')])
        print(f"  {folder}/ → {n} files")


def main():
    parser = argparse.ArgumentParser(description="Split WAV using alignment Excel timestamps")
    parser.add_argument("--xlsx", required=True, help="Path to the alignment .xlsx file")
    parser.add_argument("--wav", required=True, help="Path to the source .wav file")
    parser.add_argument("--output_dir", required=True, help="Destination folder for split audio files")
    parser.add_argument("--type", default="all",
                        choices=["all", "character", "word", "sentence", "paragraph"],
                        help="Filter by type (default: all)")
    parser.add_argument("--format", default="wav", choices=["wav", "mp3", "flac"],
                        help="Output audio format (default: wav)")
    args = parser.parse_args()

    if not Path(args.xlsx).exists():
        sys.exit(f"Error: Excel file not found: {args.xlsx}")
    if not Path(args.wav).exists():
        sys.exit(f"Error: WAV file not found: {args.wav}")

    segments = read_excel(args.xlsx, args.type)
    print(f"Found {len(segments)} segments in Excel" +
          (f" (filtered: {args.type})" if args.type != "all" else ""))

    if not segments:
        sys.exit("No matching segments found. Check your Excel file or --type filter.")

    split_audio(args.wav, segments, args.output_dir, args.format)


if __name__ == "__main__":
    main()