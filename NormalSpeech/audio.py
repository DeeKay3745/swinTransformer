"""
Audio-Text Forced Alignment Script
===================================
Aligns text from a .docx file with a .wav audio file and produces an .xlsx
with columns: Type | Content | Start Time | End Time

Requirements (install once):
    pip install python-docx openpyxl stable-ts

stable-ts (stable-whisper) wraps OpenAI Whisper and gives much better
word-level timestamps than vanilla Whisper.

Usage:
    python align_audio_text.py --docx <path.docx> --wav <path.wav> --output <output.xlsx>
    
    Optional:
        --model  tiny|base|small|medium|large  (default: base)
"""

import argparse
import re
import sys
from pathlib import Path

import stable_whisper
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ── Helpers ──────────────────────────────────────────────────────────────────

def fmt_time(seconds: float) -> str:
    """Convert seconds to hh:mm:ss.000 format."""
    if seconds < 0:
        seconds = 0
    h = int(seconds // 3600)
    m = int((seconds % 3600) // 60)
    s = seconds % 60
    return f"{h:02d}:{m:02d}:{s:06.3f}"


def clean(text: str) -> str:
    """Lowercase, strip punctuation for matching."""
    return re.sub(r'[^a-z0-9]', '', text.lower())


# ── Parse the DOCX ──────────────────────────────────────────────────────────

def parse_docx(docx_path: str) -> dict:
    """
    Extract structured content from the specific document format:
      - Characters (individual alphabets)
      - Words
      - Sentences
      - Paragraph(s)
    Returns dict with keys: characters, words, sentences, paragraphs
    """
    doc = Document(docx_path)

    # Flatten: split each paragraph on internal newlines so soft-returns
    # (Shift+Enter) become separate lines just like hard-returns.
    lines = []
    for p in doc.paragraphs:
        for sub in p.text.split('\n'):
            s = sub.strip()
            if s:
                lines.append(s)

    characters = []
    words = []
    sentences = []
    paragraphs = []

    section = None
    para_buffer = []

    i = 0
    while i < len(lines):
        line = lines[i]

        # ── Section headers ──
        if line.lower().startswith("text material") or line.lower().startswith("part-"):
            i += 1; continue
        if "alphabet" in line.lower():
            section = "alphabet"; i += 1; continue
        if "read the words" in line.lower():
            i += 1; continue
        if re.match(r'^words?\s*(\(|$)', line.lower()):
            section = "word"; i += 1; continue
        if line.lower().startswith("sentence"):
            section = "sentence"; i += 1; continue
        if line.lower().startswith("paragraph"):
            section = "paragraph"; i += 1; continue
        # Horizontal rules / dashes
        if set(line.strip()) <= {'-', '—'}:
            i += 1; continue

        # ── Parse based on current section ──
        if section == "alphabet":
            for ch in line.split():
                if len(ch) == 1 and ch.isalpha():
                    characters.append(ch.upper())

        elif section == "word":
            w = line.strip().strip(')')
            if w and not w.startswith('(') and not w.lower().startswith("take"):
                words.append(w)

        elif section == "sentence":
            # Match "(s1) text..." or "(S1) text..."
            m = re.match(r'\(s\d+\)\s*(.*)', line, re.IGNORECASE)
            if m:
                sent_text = m.group(1).strip()
                # Peek ahead: gather continuation lines that don't start a
                # new sentence label and aren't a section header
                while i + 1 < len(lines):
                    nxt = lines[i + 1]
                    if re.match(r'\(s\d+\)', nxt, re.IGNORECASE):
                        break
                    if nxt.lower().startswith("paragraph"):
                        break
                    if set(nxt.strip()) <= {'-', '—'}:
                        break
                    # It's a continuation of the current sentence
                    sent_text += ' ' + nxt.strip()
                    i += 1
                sentences.append(sent_text)
            elif line and not line.startswith('('):
                # Unlabeled continuation or standalone sentence
                if sentences:
                    sentences[-1] += ' ' + line.strip()
                else:
                    sentences.append(line.strip())

        elif section == "paragraph":
            para_buffer.append(line.strip())

        i += 1

    if para_buffer:
        paragraphs.append(' '.join(para_buffer))

    # Debug: print what was parsed
    print(f"\n  === Parsed from DOCX ===")
    print(f"  Characters ({len(characters)}): {' '.join(characters)}")
    print(f"  Words ({len(words)}): {', '.join(words[:10])}{'...' if len(words) > 10 else ''}")
    for si, s in enumerate(sentences):
        print(f"  Sentence {si+1}: {s[:90]}{'...' if len(s) > 90 else ''}")
    print(f"  Paragraphs: {len(paragraphs)}")
    print()

    return {
        'characters': characters,
        'words': words,
        'sentences': sentences,
        'paragraphs': paragraphs,
    }


# ── Align ────────────────────────────────────────────────────────────────────

def align(wav_path: str, model_name: str = "base") -> list:
    """
    Transcribe audio with stable-whisper and return a flat list of
    word-segment dicts: [{word, start, end}, ...]
    """
    print(f"Loading Whisper model '{model_name}' …")
    model = stable_whisper.load_model(model_name)

    print("Transcribing & aligning …")
    result = model.transcribe(wav_path, word_timestamps=True)

    # stable-whisper gives segments → words
    word_list = []
    for seg in result.segments:
        for w in seg.words:
            word_list.append({
                'word': w.word.strip(),
                'start': w.start,
                'end': w.end,
            })
    return word_list


def find_best_match(needle: str, word_list: list, search_start: int = 0) -> int:
    """Find the index in word_list whose cleaned text best matches needle.
    Always returns a valid index (clamped to 0..len-1). Never returns out of bounds."""
    n = len(word_list)
    if n == 0:
        return 0
    # Clamp search_start to valid range
    search_start = min(search_start, n - 1)

    needle_c = clean(needle)
    if not needle_c:
        return search_start

    # Exact match — search forward from current position
    for i in range(search_start, n):
        if clean(word_list[i]['word']) == needle_c:
            return i

    # Exact match — wrap around and search from beginning
    for i in range(0, search_start):
        if clean(word_list[i]['word']) == needle_c:
            return i

    # Fuzzy: partial match forward
    for i in range(search_start, n):
        wc = clean(word_list[i]['word'])
        if needle_c in wc or wc in needle_c:
            return i

    # Fuzzy: partial match from beginning
    for i in range(0, search_start):
        wc = clean(word_list[i]['word'])
        if needle_c in wc or wc in needle_c:
            return i

    # No match at all — return clamped search_start
    return search_start


def map_timestamps(parsed: dict, word_list: list) -> list:
    """
    Walk through parsed doc items and map each to timestamps from word_list.
    Returns a list of dicts: [{type, content, start, end}, ...]
    """
    rows = []
    n = len(word_list)
    if n == 0:
        print("WARNING: word_list is empty — cannot map timestamps.")
        return rows

    idx = 0  # current position in word_list

    # 1. Characters
    for ch in parsed['characters']:
        i = find_best_match(ch, word_list, idx)
        rows.append({
            'type': 'Character',
            'content': ch,
            'start': word_list[i]['start'],
            'end': word_list[i]['end'],
        })
        idx = min(i + 1, n - 1)

    # 2. Words
    for w in parsed['words']:
        tokens = w.replace('-', ' ').split()
        first_i = find_best_match(tokens[0], word_list, idx)
        last_i = first_i
        for t in tokens[1:]:
            last_i = find_best_match(t, word_list, min(last_i + 1, n - 1))
        rows.append({
            'type': 'Word',
            'content': w,
            'start': word_list[first_i]['start'],
            'end': word_list[last_i]['end'],
        })
        idx = min(last_i + 1, n - 1)

    # 3. Sentences
    for sent in parsed['sentences']:
        sent_words = sent.split()
        if not sent_words:
            continue
        first_i = find_best_match(sent_words[0], word_list, idx)
        last_i = first_i
        for sw in sent_words[1:]:
            last_i = find_best_match(sw, word_list, last_i)
        rows.append({
            'type': 'Sentence',
            'content': sent,
            'start': word_list[first_i]['start'],
            'end': word_list[last_i]['end'],
        })
        idx = min(last_i + 1, n - 1)

    # 4. Paragraphs
    for para in parsed['paragraphs']:
        para_words = para.split()
        if not para_words:
            continue
        first_i = find_best_match(para_words[0], word_list, idx)
        last_i = first_i
        for pw in para_words[1:]:
            last_i = find_best_match(pw, word_list, last_i)
        rows.append({
            'type': 'Paragraph',
            'content': para[:80] + ('…' if len(para) > 80 else ''),
            'start': word_list[first_i]['start'],
            'end': word_list[last_i]['end'],
        })
        idx = min(last_i + 1, n - 1)

    return rows


# ── Excel Output ─────────────────────────────────────────────────────────────

def write_excel(rows: list, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Alignment"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["Type", "Content", "Start Time (hh:mm:ss.000)", "End Time (hh:mm:ss.000)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Type-based color coding
    type_colors = {
        'Character': 'E2EFDA',
        'Word': 'DDEBF7',
        'Sentence': 'FCE4D6',
        'Paragraph': 'E4DFEC',
    }
    data_font = Font(name="Arial", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for r, row in enumerate(rows, 2):
        fill = PatternFill("solid", fgColor=type_colors.get(row['type'], 'FFFFFF'))

        c1 = ws.cell(row=r, column=1, value=row['type'])
        c1.font = data_font; c1.fill = fill; c1.alignment = center_align; c1.border = thin_border

        c2 = ws.cell(row=r, column=2, value=row['content'])
        c2.font = data_font; c2.fill = fill; c2.alignment = left_align; c2.border = thin_border

        c3 = ws.cell(row=r, column=3, value=fmt_time(row['start']))
        c3.font = data_font; c3.fill = fill; c3.alignment = center_align; c3.border = thin_border

        c4 = ws.cell(row=r, column=4, value=fmt_time(row['end']))
        c4.font = data_font; c4.fill = fill; c4.alignment = center_align; c4.border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 26
    ws.column_dimensions['D'].width = 26

    # Freeze top row
    ws.freeze_panes = 'A2'

    wb.save(output_path)
    print(f"Excel saved → {output_path}")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Align text from a .docx with audio from a .wav → .xlsx timestamps"
    )
    parser.add_argument("--docx", required=True, help="Path to the .docx file")
    parser.add_argument("--wav", required=True, help="Path to the .wav audio file")
    parser.add_argument("--output", default="alignment_output.xlsx", help="Output .xlsx path")
    parser.add_argument("--model", default="base", choices=["tiny", "base", "small", "medium", "large"],
                        help="Whisper model size (default: base)")
    args = parser.parse_args()

    # Validate inputs
    if not Path(args.docx).exists():
        sys.exit(f"Error: DOCX file not found: {args.docx}")
    if not Path(args.wav).exists():
        sys.exit(f"Error: WAV file not found: {args.wav}")

    # Step 1: Parse the document
    print("Parsing DOCX …")
    parsed = parse_docx(args.docx)
    print(f"  Found: {len(parsed['characters'])} characters, "
          f"{len(parsed['words'])} words, "
          f"{len(parsed['sentences'])} sentences, "
          f"{len(parsed['paragraphs'])} paragraph(s)")

    # Step 2: Transcribe + get word-level timestamps
    word_list = align(args.wav, args.model)
    print(f"  Whisper detected {len(word_list)} word-tokens in audio")

    if not word_list:
        sys.exit("Error: No words detected in audio. Check your WAV file.")

    # Debug: show first 30 detected words
    print("\n  First 30 detected words from audio:")
    for i, ww in enumerate(word_list[:30]):
        print(f"    [{i:3d}] {ww['word']:20s}  {fmt_time(ww['start'])} → {fmt_time(ww['end'])}")
    print()

    # Step 3: Map document items to timestamps
    print("Mapping timestamps …")
    rows = map_timestamps(parsed, word_list)

    # Step 4: Write Excel
    write_excel(rows, args.output)
    print("Done!")


if __name__ == "__main__":
    main()