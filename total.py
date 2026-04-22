"""
WAV Duration Report Generator — Mixed Structure Support
Handles two layouts automatically:
  • Root / FolderA / *.wav              (direct WAVs)
  • Root / FolderB / SubFolder / *.wav  (nested WAVs)

Supports ALL WAV variants:
  • Standard PCM RIFF WAV   → read by built-in `wave`
  • RF64 / WAVE64 / BWF     → read by `soundfile`  (pip install soundfile)
  • Any other audio format  → read by `pydub`      (pip install pydub + ffmpeg)

Usage:
    python wav_duration_report.py /path/to/root/folder
    python wav_duration_report.py /path/to/root/folder --output report.xlsx

Requirements:
    pip install openpyxl soundfile pydub
    # pydub also needs ffmpeg:  https://ffmpeg.org/download.html
"""

import sys
import wave
import argparse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Optional backends — imported lazily so missing ones just skip silently
try:
    import soundfile as _sf
    _HAS_SOUNDFILE = True
except ImportError:
    _HAS_SOUNDFILE = False

try:
    from pydub import AudioSegment as _AS
    _HAS_PYDUB = True
except ImportError:
    _HAS_PYDUB = False


# ── Helpers ──────────────────────────────────────────────────────────────────

def wav_duration_seconds(filepath: Path) -> float:
    """
    Try three backends in order:
      1. wave      — fast, zero deps, standard RIFF only
      2. soundfile — handles RF64 / WAVE64 / BWF / float WAV
      3. pydub     — handles virtually anything (needs ffmpeg)
    Returns 0.0 and prints a warning only if ALL backends fail.
    """
    # ── Backend 1: built-in wave ──────────────────────────────────────────
    try:
        with wave.open(str(filepath), "rb") as wf:
            return wf.getnframes() / float(wf.getframerate())
    except Exception:
        pass

    # ── Backend 2: soundfile (RF64 / WAVE64 / BWF etc.) ──────────────────
    if _HAS_SOUNDFILE:
        try:
            info = _sf.info(str(filepath))
            return info.frames / float(info.samplerate)
        except Exception:
            pass

    # ── Backend 3: pydub (anything ffmpeg can open) ───────────────────────
    if _HAS_PYDUB:
        try:
            audio = _AS.from_file(str(filepath))
            return len(audio) / 1000.0      # pydub uses milliseconds
        except Exception:
            pass

    # ── All backends failed ───────────────────────────────────────────────
    missing = []
    if not _HAS_SOUNDFILE: missing.append("soundfile")
    if not _HAS_PYDUB:     missing.append("pydub")
    hint = (f"  → Install missing backends: pip install {' '.join(missing)}"
            if missing else "")
    print(f"  [WARN] Could not read {filepath.name} — all backends failed.{hint}")
    return 0.0


def fmt(seconds: float) -> str:
    """Format seconds -> hh:mm:ss.000"""
    ms = round(seconds * 1000)
    s  = ms // 1000;  ms %= 1000
    m  = s  // 60;    s  %= 60
    h  = m  // 60;    m  %= 60
    return f"{h:02d}:{m:02d}:{s:02d}.{ms:03d}"


def find_wavs_direct(folder: Path) -> list[Path]:
    """Return WAV files sitting directly inside folder (not recursive)."""
    return sorted(
        p for p in folder.iterdir()
        if p.is_file() and p.suffix.lower() == ".wav"
    )


# ── Scanner ───────────────────────────────────────────────────────────────────
#
# Builds a list of top-level folder records.
# Each record has one or more "groups":
#   - subfolder=None   -> WAVs found directly in the top folder
#   - subfolder="Name" -> WAVs found inside a named sub-folder
#
# Structure returned:
# [
#   {
#     "name":    "FolderA",
#     "groups": [
#       {"subfolder": None,       "files": [...], "total_s": 12.3},
#       {"subfolder": "SubFolderX", "files": [...], "total_s":  5.1},
#     ],
#     "total_s": 17.4,
#   },
#   ...
# ]

def scan(root: Path) -> list[dict]:
    records = []

    for top in sorted(p for p in root.iterdir() if p.is_dir()):
        groups = []

        # Case 1: WAVs directly inside this top-level folder
        direct = find_wavs_direct(top)
        if direct:
            files = [{"name": w.name, "duration_s": wav_duration_seconds(w)} for w in direct]
            groups.append({
                "subfolder": None,
                "files":     files,
                "total_s":   sum(f["duration_s"] for f in files),
            })

        # Case 2: Sub-folders containing WAVs
        for sub in sorted(p for p in top.iterdir() if p.is_dir()):
            sub_wavs = find_wavs_direct(sub)
            if not sub_wavs:
                continue
            files = [{"name": w.name, "duration_s": wav_duration_seconds(w)} for w in sub_wavs]
            groups.append({
                "subfolder": sub.name,
                "files":     files,
                "total_s":   sum(f["duration_s"] for f in files),
            })

        if groups:
            records.append({
                "name":    top.name,
                "groups":  groups,
                "total_s": sum(g["total_s"] for g in groups),
            })

    return records


# ── Excel builder ─────────────────────────────────────────────────────────────

def build_excel(records: list[dict], out: Path):

    C_NAVY  = "1F4E79"
    C_BLUE  = "2E75B6"
    C_LBLUE = "D6E4F0"
    C_GOLD  = "FFC000"
    C_LGOLD = "FFF2CC"
    C_WHITE = "FFFFFF"
    C_GREY  = "F2F2F2"

    def mkfill(hex_):
        return PatternFill("solid", start_color=hex_)

    def mkfont(bold=False, color=C_NAVY, size=10):
        return Font(name="Arial", bold=bold, color=color, size=size)

    def mkalign(h="left"):
        return Alignment(horizontal=h, vertical="center")

    thin  = Side(style="thin",   color="B8CCE4")
    thick = Side(style="medium", color=C_NAVY)
    brd       = Border(left=thin, right=thin, top=thin, bottom=thin)
    brd_thick = Border(left=thin, right=thin, top=thin, bottom=thick)

    def style(cell, *, bold=False, color=C_NAVY, size=10, bg=None,
              h="left", thick_bottom=False):
        cell.font      = mkfont(bold, color, size)
        cell.alignment = mkalign(h)
        cell.border    = brd_thick if thick_bottom else brd
        if bg:
            cell.fill = mkfill(bg)

    wb = Workbook()

    # ── SUMMARY sheet ─────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 22

    for col, title in enumerate(["Folder", "WAV Files", "Total Duration"], 1):
        c = ws.cell(1, col, title)
        style(c, bold=True, color=C_WHITE, size=11, bg=C_NAVY, h="center")

    grand_files = 0
    grand_s     = 0.0

    for i, rec in enumerate(records):
        r   = i + 2
        fc  = sum(len(g["files"]) for g in rec["groups"])
        grand_files += fc
        grand_s     += rec["total_s"]
        bg = C_GREY if i % 2 else C_WHITE

        style(ws.cell(r, 1, rec["name"]),         bg=bg)
        style(ws.cell(r, 2, fc),                  bg=bg, h="center")
        style(ws.cell(r, 3, fmt(rec["total_s"])), bg=bg, h="center")
        ws.row_dimensions[r].height = 16

    tr = len(records) + 2
    ws.row_dimensions[tr].height = 22
    for col, val in enumerate([("GRAND TOTAL", "center"),
                                (grand_files,   "center"),
                                (fmt(grand_s),  "center")], 1):
        c = ws.cell(tr, col, val[0])
        style(c, bold=True, size=11, bg=C_GOLD, h=val[1], thick_bottom=True)

    # ── DETAILS sheet ─────────────────────────────────────────────────────────
    wd = wb.create_sheet("Details")
    wd.freeze_panes = "A2"
    wd.row_dimensions[1].height = 22
    wd.column_dimensions["A"].width = 30
    wd.column_dimensions["B"].width = 30
    wd.column_dimensions["C"].width = 42
    wd.column_dimensions["D"].width = 22

    for col, title in enumerate(
        ["Top Folder", "Sub-Folder", "File Name", "Duration"], 1
    ):
        c = wd.cell(1, col, title)
        style(c, bold=True, color=C_WHITE, size=11, bg=C_NAVY, h="center")

    dr = 2
    for rec in records:
        for grp in rec["groups"]:
            sub_label = grp["subfolder"] or "—"

            for file in grp["files"]:
                bg = C_GREY if dr % 2 == 0 else C_WHITE
                style(wd.cell(dr, 1, rec["name"]),           bg=bg)
                style(wd.cell(dr, 2, sub_label),             bg=bg)
                style(wd.cell(dr, 3, file["name"]),          bg=bg)
                style(wd.cell(dr, 4, fmt(file["duration_s"])), bg=bg, h="center")
                wd.row_dimensions[dr].height = 15
                dr += 1

            # Sub-folder subtotal row (only when it is a named sub-folder)
            if grp["subfolder"] is not None:
                wd.row_dimensions[dr].height = 17
                style(wd.cell(dr, 1, rec["name"]),
                      bold=True, color=C_BLUE, bg=C_LBLUE)
                style(wd.cell(dr, 2, f"Subtotal — {grp['subfolder']}"),
                      bold=True, color=C_BLUE, bg=C_LBLUE)
                style(wd.cell(dr, 3, f"{len(grp['files'])} file(s)"),
                      bold=True, color=C_BLUE, bg=C_LBLUE, h="center")
                style(wd.cell(dr, 4, fmt(grp["total_s"])),
                      bold=True, color=C_BLUE, bg=C_LBLUE, h="center")
                dr += 1

        # Top-folder total row
        wd.row_dimensions[dr].height = 18
        total_files = sum(len(g["files"]) for g in rec["groups"])
        style(wd.cell(dr, 1, f"Total — {rec['name']}"),
              bold=True, size=11, bg=C_LGOLD, thick_bottom=True)
        style(wd.cell(dr, 2, ""),
              bg=C_LGOLD, thick_bottom=True)
        style(wd.cell(dr, 3, f"{total_files} file(s)"),
              bold=True, size=11, bg=C_LGOLD, h="center", thick_bottom=True)
        style(wd.cell(dr, 4, fmt(rec["total_s"])),
              bold=True, size=11, bg=C_LGOLD, h="center", thick_bottom=True)
        dr += 1

    # Grand total row
    wd.row_dimensions[dr].height = 22
    all_files = sum(len(g["files"]) for r in records for g in r["groups"])
    all_s     = sum(r["total_s"] for r in records)
    for col, val in enumerate([
        ("GRAND TOTAL",      "center"),
        ("",                 "center"),
        (f"{all_files} file(s)", "center"),
        (fmt(all_s),         "center"),
    ], 1):
        c = wd.cell(dr, col, val[0])
        style(c, bold=True, size=12, bg=C_GOLD, h=val[1], thick_bottom=True)

    wb.save(out)
    print(f"\n✅  Saved: {out}")


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="WAV duration report — handles flat and nested folder layouts."
    )
    parser.add_argument("root",
                        help="Root folder containing subfolders with WAV files")
    parser.add_argument("--output", default="wav_duration_report.xlsx",
                        help="Output Excel filename (default: wav_duration_report.xlsx)")
    args = parser.parse_args()

    root = Path(args.root).expanduser().resolve()
    if not root.is_dir():
        print(f"❌  Not a valid directory: {root}")
        sys.exit(1)

    print(f"📂  Scanning: {root}")
    records = scan(root)

    if not records:
        print("⚠️  No WAV files found.")
        sys.exit(0)

    total_files = sum(len(g["files"]) for r in records for g in r["groups"])
    total_s     = sum(r["total_s"] for r in records)
    print(f"   {len(records)} folder(s)  |  {total_files} WAV file(s)  |  {fmt(total_s)}")

    build_excel(records, Path(args.output).expanduser().resolve())


if __name__ == "__main__":
    main()